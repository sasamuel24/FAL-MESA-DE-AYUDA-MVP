"""
Servicio de conversión directa Excel-PDF manteniendo formato original
Usa LibreOffice headless para conversión nativa sin reconstruir el diseño
"""

import os
import shutil
import subprocess
import tempfile
import logging
import requests
import base64
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from copy import copy
from datetime import datetime
from PIL import Image as PILImage
import io

logger = logging.getLogger(__name__)

class ExcelDirectPDFService:
    """
    Servicio que mantiene el diseño exacto del Excel original
    Solo llena campos y convierte con LibreOffice headless
    """
    
    def __init__(self, template_path: Optional[str] = None, s3_service=None, db_session=None):
        """
        Inicializar servicio con ruta de plantilla
        
        Args:
            template_path: Ruta al archivo Excel plantilla
            s3_service: Servicio S3 para descargar imágenes
            db_session: Sesión de base de datos para obtener firmas
        """
        if template_path:
            self.template_path = Path(template_path)
        else:
            # Ruta por defecto al template en el proyecto
            self.template_path = Path(__file__).parent.parent / "templates" / "FO-MT-006 Orden de trabajo de mantenimiento v1.xlsx"
        
        if not self.template_path.exists():
            raise FileNotFoundError(f"Plantilla Excel no encontrada: {self.template_path}")
        
        # Importar S3Service si no se proporciona
        if s3_service is None:
            try:
                from app.services.s3_service import S3Service
                self.s3_service = S3Service()
            except ImportError:
                logger.warning("S3Service no disponible, las imágenes no se podrán insertar")
                self.s3_service = None
        else:
            self.s3_service = s3_service
        
        # Guardar sesión de base de datos
        self.db_session = db_session
        
        logger.info(f"ExcelDirectPDFService inicializado con plantilla: {self.template_path}")
    
    def generar_pdf_directo(self, datos_ot: Dict[str, Any], folio: int) -> bytes:
        """
        Generar PDF manteniendo diseño exacto del Excel original
        
        Args:
            datos_ot: Diccionario con datos de la OT
            folio: Número de folio de la OT
            
        Returns:
            bytes: Contenido del PDF generado
        """
        logger.info(f"Generando PDF directo para OT {folio} manteniendo diseño original")
        
        # Crear directorio temporal para trabajar
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            
            # PASO 1: Copiar plantilla original a directorio temporal
            excel_temp = temp_path / f"OT_{folio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            shutil.copy2(self.template_path, excel_temp)
            logger.info(f"Plantilla copiada a: {excel_temp}")
            
            # PASO 2: Llenar solo los campos necesarios SIN modificar formato
            self._llenar_campos_excel(excel_temp, datos_ot, folio)
            
            # PASO 3: Intentar conversión con LibreOffice primero, Excel COM como fallback
            pdf_path = None
            
            if self.verificar_libreoffice():
                try:
                    pdf_path = self._convertir_con_libreoffice(excel_temp, temp_path)
                except Exception as e:
                    logger.warning(f"LibreOffice falló, intentando con Excel COM: {e}")
                    pdf_path = None
            
            # Fallback: Excel COM en Windows (solo para desarrollo)
            if pdf_path is None and os.name == 'nt':
                try:
                    pdf_path = self._convertir_con_excel_com(excel_temp, temp_path)
                except Exception as e:
                    logger.error(f"Excel COM también falló: {e}")
                    raise RuntimeError("No se pudo convertir con LibreOffice ni Excel COM")
            
            if pdf_path is None:
                raise RuntimeError("LibreOffice no disponible y no es sistema Windows para Excel COM")
            
            # PASO 4: Leer el PDF generado y retornarlo como bytes
            with open(pdf_path, 'rb') as pdf_file:
                pdf_content = pdf_file.read()
            
            logger.info(f"PDF generado exitosamente para OT {folio}")
            return pdf_content
    
    def _llenar_campos_excel(self, excel_path: Path, datos_ot: Dict[str, Any], folio: int):
        """
        Llenar campos específicos del Excel usando la estructura JSON especificada
        Maneja celdas combinadas y mapea correctamente según los requerimientos
        
        Args:
            excel_path: Ruta al archivo Excel temporal
            datos_ot: Datos de la OT en formato JSON según especificación
            folio: Número de folio
        """
        logger.info("Llenando campos en Excel manteniendo formato original...")
        logger.info(f"Datos JSON recibidos para OT {folio}: {list(datos_ot.keys())}")
        
        # Log específico para notas del técnico
        notas_tecnico = datos_ot.get('notas_tecnico', '')
        if notas_tecnico:
            logger.info(f"📝 Notas del técnico encontradas: '{notas_tecnico[:100]}{'...' if len(notas_tecnico) > 100 else ''}'")
        else:
            logger.info("📝 No se encontraron notas del técnico")
        
        # Cargar workbook manteniendo formato y estilos
        wb = load_workbook(excel_path, data_only=False)
        ws = wb.active
        
        # Función auxiliar para manejar celdas combinadas
        def escribir_celda_segura(celda_coord: str, valor: str):
            """Escribe en una celda manejando casos de celdas combinadas"""
            try:
                if not valor:
                    logger.warning(f"⚠️ Celda {celda_coord} tiene valor vacío, saltando...")
                    return False
                    
                cell = ws[celda_coord]
                
                # Verificar si la celda está en un rango combinado
                cell_in_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Usar la celda superior izquierda del rango combinado
                        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = str(valor)
                        logger.info(f"✅ Campo llenado en rango combinado: {celda_coord} (usando {top_left_cell.coordinate}) = '{valor}'")
                        cell_in_merged = True
                        return True
                
                if not cell_in_merged:
                    # Celda normal, no combinada
                    cell.value = str(valor)
                    logger.info(f"✅ Campo llenado: {celda_coord} = '{valor}'")
                    return True
                    
            except Exception as e:
                logger.error(f"❌ Error escribiendo celda {celda_coord}: {str(e)}")
                return False

        # Función auxiliar para configurar texto con ajuste automático
        def configurar_celda_texto_largo(celda_coord: str, valor: str, alto_fila: int = None):
            """
            Configura una celda para manejar texto largo con ajuste automático
            """
            try:
                if not valor:
                    logger.info(f"⚠️ No hay contenido para celda {celda_coord}")
                    return False
                
                cell = ws[celda_coord]
                
                # Verificar si está en rango combinado y usar celda principal
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        break
                
                # Establecer el valor
                cell.value = str(valor)
                
                # Activar ajuste de texto (wrap text)
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='top',
                    horizontal='left'
                )
                
                # Configurar altura de fila si se especifica
                if alto_fila:
                    ws.row_dimensions[cell.row].height = alto_fila
                
                logger.info(f"✅ Celda {celda_coord} configurada con ajuste de texto")
                logger.info(f"📝 Contenido: '{valor[:100]}{'...' if len(valor) > 100 else ''}'")
                return True
                
            except Exception as e:
                logger.error(f"❌ Error configurando celda {celda_coord}: {str(e)}")
                return False

        # Función auxiliar para manejar notas del técnico en área grande (A35:H36)
        def escribir_notas_tecnico(notas: str):
            try:
                if not notas:
                    logger.info("⚠️ No hay notas del técnico para escribir")
                    return False
                
                # Usar la función de configuración avanzada para las notas
                # El rango A35:H36 es combinado, así que A35 es la celda principal
                return configurar_celda_texto_largo('A35', notas, alto_fila=60)
                
            except Exception as e:
                logger.error(f"❌ Error escribiendo notas del técnico: {str(e)}")
                return False
        
        # Función auxiliar para manejar descripción del trabajo (A25:H28)
        def escribir_descripcion_trabajo(descripcion: str):
            try:
                if not descripcion:
                    logger.info("⚠️ No hay descripción del trabajo para escribir")
                    return False
                
                # Usar la función de configuración avanzada para la descripción
                # El rango A25:H28 es combinado, así que A25 es la celda principal
                return configurar_celda_texto_largo('A25', descripcion, alto_fila=80)
                
            except Exception as e:
                logger.error(f"❌ Error escribiendo descripción del trabajo: {str(e)}")
                return False
        
        # Mapeo exacto según el formato JSON y la plantilla Excel
        campos_excel = {
            # Encabezado - Información básica
            'G5': datos_ot.get('id', str(folio)),  # ID (folio)
            'B7': datos_ot.get('fecha', datetime.now().strftime('%d/%m/%Y')),  # Fecha de exportación
            'Q4': f"FO-MT-006-{datos_ot.get('id', str(folio))}",  # Código del documento
            'Q5': '1.0',  # Versión
            
            # Título y estado - Manejo especial para celdas combinadas
            'B5': datos_ot.get('titulo', f'OT #{folio}'),  # Título 
            'F7': datos_ot.get('estado', 'En proceso'),  # Estado
            
            # Información de la OT - sección Media
            'C11': datos_ot.get('categoria', ''),  # Categoría
            'G11': datos_ot.get('subcategoria', ''),  # Subcategoría
            'C12': datos_ot.get('zona_planta_tienda', ''),  # Zona/Planta/Tienda
            'G12': datos_ot.get('ciudad', ''),  # Ciudad
            'C13': datos_ot.get('prioridad', ''),  # Prioridad
            'G13': datos_ot.get('tipo_solicitud', ''),  # Tipo de Solicitud
            'C14': datos_ot.get('tipo_mantenimiento', ''),  # Tipo de Mantenimiento
            'G14': datos_ot.get('tiempo_estimado_horas', ''),  # Tiempo estimado
            'C15': datos_ot.get('etapa', ''),  # Etapa
            
            # Asignación
            'C19': datos_ot.get('asignacion', {}).get('tecnico_asignado', '') or 'Por asignar',  # Técnico asignado
            'G19': datos_ot.get('asignacion', {}).get('fecha_visita', '') or '',  # Fecha de visita
            'C20': datos_ot.get('solicitante', ''),  # Solicitante
            # 'G20': se manejará por configurar_contacto_solicitante para adaptarse a la celda
            
            # Descripción y notas serán manejadas por funciones especializadas
            # 'A25': datos_ot.get('descripcion', ''),  # Descripción del trabajo - MANEJADO POR escribir_descripcion_trabajo
            # 'A35': datos_ot.get('notas_tecnico', ''),  # Notas del técnico - MANEJADO POR escribir_notas_tecnico
            
            # Archivos adjuntos - Solo mostrar texto si no hay imágenes o hay otros archivos
            # 'A41': se manejará condicionalmente después
            
            # Firmas de conformidad - POSICIÓN DINÁMICA SEGÚN ÁREA DE IMÁGENES
            # Las firmas se reposicionarán automáticamente después del análisis de imágenes
        }
        
        logger.info(f"Iniciando llenado de {len(campos_excel)} campos...")
        
        campos_completados = 0
        
        # Procesar campos regulares primero
        for celda_coord, valor in campos_excel.items():
            logger.info(f"Procesando celda {celda_coord} con valor: '{valor}'")
            if escribir_celda_segura(celda_coord, valor):
                campos_completados += 1
        
        # Manejar archivos adjuntos condicionalmente
        texto_archivos = self._generar_texto_archivos_adjuntos(datos_ot.get('archivos_adjuntos', []))
        if texto_archivos:  # Solo escribir si hay texto (archivos no-imagen)
            if escribir_celda_segura('A41', texto_archivos):
                campos_completados += 1
        
        # Manejar campos especiales con ajuste de texto
        
        # 1. Descripción del trabajo (A25:H28)
        descripcion = datos_ot.get('descripcion', '')
        if descripcion:
            if escribir_descripcion_trabajo(descripcion):
                campos_completados += 1
                logger.info("✅ Descripción del trabajo procesada con ajuste de texto")
        
        # 2. Notas del técnico (A35:H36) 
        notas_tecnico = datos_ot.get('notas_tecnico', '')
        if notas_tecnico:
            if escribir_notas_tecnico(notas_tecnico):
                campos_completados += 1
                logger.info("✅ Notas del técnico procesadas con ajuste de texto")
        
        # 3. Contacto solicitante (G20) con adaptación automática
        contacto_solicitante = datos_ot.get('contacto_solicitante', '')
        if contacto_solicitante:
            if self._configurar_contacto_solicitante(ws, contacto_solicitante):
                campos_completados += 1
                logger.info("✅ Contacto solicitante configurado con adaptación automática")
        
        # Insertar imagen original de la solicitud (área A30:A32)
        self._insertar_imagen_original(ws, datos_ot.get('imagen_original', ''))
        
        # Insertar imágenes de archivos adjuntos con validación anti-solapamiento
        area_expandida = self._insertar_imagenes_adjuntas_seguro(ws, datos_ot.get('archivos_adjuntos', []))
        
        # Reposicionar firmas según el área expandida utilizada (con validación estricta)
        self._reposicionar_firmas_dinamicamente(ws, area_expandida, datos_ot)
        
        # ⚡ CONFIGURAR ALTURAS DE ÁREAS DE FIRMA PARA LIBREOFFICE - POSICIÓN FIJA
        self._configurar_alturas_firmas_libreoffice(ws, 48)
        
        # Guardar manteniendo formato original
        wb.save(excel_path)
        logger.info(f"Excel llenado: {campos_completados}/{len(campos_excel)} campos completados exitosamente")
        
        # Log de información adicional para debugging
        if datos_ot.get('archivos_adjuntos'):
            logger.info(f"Archivos adjuntos encontrados: {len(datos_ot['archivos_adjuntos'])}")
            for i, archivo in enumerate(datos_ot['archivos_adjuntos']):
                logger.info(f"  Archivo {i+1}: {archivo.get('nombre', 'Sin nombre')}")
    
    def _generar_texto_archivos_adjuntos(self, archivos_adjuntos: list) -> str:
        """
        Generar texto descriptivo para archivos adjuntos
        Las imágenes se mostrarán visualmente sin texto descriptivo
        """
        if not archivos_adjuntos:
            return 'Sin archivos adjuntos'
        
        # Separar imágenes de otros archivos
        imagenes = []
        otros_archivos = []
        
        for archivo in archivos_adjuntos:
            nombre = archivo.get('nombre', '')
            if nombre and any(ext in nombre.lower() for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']):
                imagenes.append(archivo)
            else:
                otros_archivos.append(archivo)
        
        texto_partes = []
        
        # Solo mostrar texto para archivos que NO son imágenes
        # Las imágenes se mostrarán visualmente sin texto descriptivo
        if otros_archivos:
            base_url = os.getenv('BASE_URL', 'http://localhost:8000') + "/api/v1/work-orders"
            
            if len(otros_archivos) == 1:
                archivo = otros_archivos[0]
                nombre = archivo.get('nombre', 'Sin nombre')
                enlace = f"{base_url}/descargar-archivo/{nombre}"
                texto_partes.append(f"Archivo adjunto: {nombre}")
                texto_partes.append(f"Descargar: {enlace}")
            else:
                texto_partes.append(f"Archivos adjuntos ({len(otros_archivos)}):")
                for i, archivo in enumerate(otros_archivos, 1):
                    nombre = archivo.get('nombre', f'Archivo_{i}')
                    enlace = f"{base_url}/descargar-archivo/{nombre}"
                    texto_partes.append(f"  {i}. {nombre}")
                    texto_partes.append(f"     Descargar: {enlace}")
        
        # Si solo hay imágenes, no mostrar texto
        if imagenes and not otros_archivos:
            return ""  # Texto vacío, solo se mostrarán las imágenes visualmente
        
        return "\n".join(texto_partes) if texto_partes else 'Sin archivos adjuntos'
    
    def _reposicionar_firmas_dinamicamente(self, ws, area_expandida: dict, datos_ot: dict):
        """
        🔄 Posicionar las firmas de conformidad en posiciones fijas: A47 (técnico) y E47 (cliente)
        POSICIÓN FIJA PARA MANTENER CONSISTENCIA EN EL FORMATO
        """
        try:
            # Nota: Los parámetros de área expandida ya no se usan, mantenemos por compatibilidad
            # Las firmas ahora van en posición fija A47 y E47
            
            # 🖊️ Obtener datos reales de firmas de conformidad desde BD
            ot_id_str = datos_ot.get('id')
            ot_id = int(ot_id_str) if ot_id_str and ot_id_str.isdigit() else None
            logger.info(f"🔍 PDF Service: Buscando firmas para OT ID {ot_id} (recibido como '{ot_id_str}')")
            firmas_data = self._obtener_firmas_conformidad(ot_id) if ot_id else {}
            logger.info(f"📋 Datos de firmas obtenidos: {firmas_data}")
            
            # Usar nombres reales de las firmas o valores por defecto
            tecnico_asignado = firmas_data.get('nombre_tecnico', '') or datos_ot.get('asignacion', {}).get('tecnico_asignado', '') or 'Por asignar'
            solicitante = firmas_data.get('nombre_cliente', '') or datos_ot.get('solicitante', '') or 'Cliente'
            
            # NOTA: Código legacy completamente removido - ahora usamos posición fija A47 y E47
            
            # 🎯 USAR POSICIÓN FIJA: A45 (nombre técnico) y E47 (nombre cliente)
            fila_nombre_tecnico = 45
            fila_nombre_cliente = 45
            
            logger.info(f"🎯 Usando posiciones fijas para nombres:")
            logger.info(f"   👨‍🔧 Técnico: '{tecnico_asignado}' → A{fila_nombre_tecnico}")
            logger.info(f"   👤 Cliente: '{solicitante}' → E{fila_nombre_cliente}")
            
            # Escribir nombres en posiciones específicas: A45 (técnico) y E47 (cliente)
            firmas_coords = {
                f'A{fila_nombre_tecnico}': tecnico_asignado,  # Nombre del Técnico en A45
                f'E{fila_nombre_cliente}': solicitante,       # Nombre del Cliente en E47
            }
            
            logger.info(f"📝 Escribiendo nombres: A{fila_nombre_tecnico} (técnico) y E{fila_nombre_cliente} (cliente)")
            for celda_coord, valor in firmas_coords.items():
                try:
                    cell = ws[celda_coord]
                    
                    # Verificar si la celda está en un rango combinado
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Usar la celda principal del rango combinado
                            cell = ws[merged_range.coord.split(':')[0]]
                            break
                    
                    cell.value = str(valor)
                    
                    # 🎯 CONFIGURAR ALINEAMIENTO CENTRADO PARA LOS NOMBRES
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    logger.info(f"✅ Firma escrita: {celda_coord} = '{valor}' (centrada)")
                    
                except Exception as e:
                    logger.error(f"❌ Error escribiendo firma {celda_coord}: {str(e)}")
            
            # 🎨 Insertar imágenes de firma (si existen) - usando posición fija
            logger.info(f"🔍 DEBUG: Datos de firmas para insertar imágenes: {firmas_data}")
            self._insertar_imagenes_firmas(ws, firmas_data, 48)  # Las imágenes van en fila 48
            
            # ⚡ Configurar alturas para nombres (A45, E45) e imágenes (A48, E48)  
            ws.row_dimensions[45].height = 20  # Altura para nombres técnico (A45) y cliente (E45)
            
            # ⚡ Configurar alturas específicas para las firmas en posición fija (fila 48)
            self._configurar_alturas_firmas_libreoffice(ws, 48)
                    
        except Exception as e:
            logger.error(f"❌ Error crítico reposicionando firmas: {str(e)}")
    
    def _insertar_imagenes_adjuntas_seguro(self, ws, archivos_adjuntos: list) -> dict:
        """
        🛡️ Insertar imágenes con validación anti-solapamiento estricta
        Versión SIMPLIFICADA pero SEGURA para resolver el problema inmediatamente
        
        Returns:
            dict: Información del área utilizada con validación anti-solapamiento
        """
        try:
            if not archivos_adjuntos or not self.s3_service:
                logger.info("⚠️ No hay archivos adjuntos para insertar o S3Service no disponible")
                return {'expandido': False, 'nueva_fila_firmas': 43, 'area_utilizada': (40, 42)}
            
            # Filtrar solo archivos de imagen y limitar a máximo 3 para mantener estilo PDF
            imagenes = []
            for archivo in archivos_adjuntos:
                nombre = archivo.get('nombre', '')
                if nombre and any(ext in nombre.lower() for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']):
                    imagenes.append(archivo)
                    # 🔢 LIMITAR A MÁXIMO 3 IMÁGENES PARA PRESERVAR ESTILO PDF
                    if len(imagenes) >= 3:
                        logger.info(f"📊 Limitando a 3 imágenes máximo para mantener estilo PDF consistente")
                        break
            
            if not imagenes:
                logger.info("⚠️ No se encontraron imágenes para insertar")
                return {'expandido': False, 'nueva_fila_firmas': 43, 'area_utilizada': (40, 42)}
            
            logger.info(f"🛡️ INSERCIÓN SEGURA: Procesando {len(imagenes)} imagen(es)")
            
            # 🔢 CALCULAR ÁREA REQUERIDA DE FORMA SIMPLE PERO SEGURA
            num_imagenes = len(imagenes)
            filas_necesarias = 8  # Base mínima para imágenes
            
            # Ajuste dinámico según cantidad de imágenes
            if num_imagenes > 2:
                filas_necesarias = 10  # Más espacio para múltiples imágenes
            if num_imagenes > 4:
                filas_necesarias = 12  # Espacio extra para muchas imágenes
            
            # 🎯 CALCULAR POSICIÓN FINAL DE ÁREA DE IMÁGENES
            area_inicio = 40  # Fila de inicio correcta para archivos adjuntos (A40:H42)
            area_fin = area_inicio + filas_necesarias - 1
            
            # 🚨 VALIDACIÓN CRÍTICA ANTI-SOLAPAMIENTO
            # Las firmas deben estar MÍNIMO 3 filas después del área de imágenes
            nueva_fila_firmas_calculada = area_fin + 4  # 3 filas de separación + 1 para ser seguro
            
            logger.info(f"🛡️ Área de imágenes calculada: filas {area_inicio}-{area_fin}")
            logger.info(f"🛡️ Nueva posición firmas (SEGURA): fila {nueva_fila_firmas_calculada}")
            
            # 📏 INSERTAR IMÁGENES CON DIMENSIONES FIJAS SEGURAS
            imagen_width, imagen_height = 200, 150  # Dimensiones conservadoras
            col_inicio = 2  # Columna B
            max_cols = 3   # Máximo 3 columnas
            
            for i, archivo in enumerate(imagenes[:6]):  # Máximo 6 imágenes para evitar overflow
                try:
                    nombre = archivo.get('nombre', '')
                    logger.info(f"🎨 Insertando imagen {i+1}: {nombre}")
                    
                    # Descargar imagen
                    url_s3 = archivo.get('url', '')
                    imagen_bytes = self._descargar_imagen_desde_url_completa(url_s3)
                    if not imagen_bytes:
                        continue
                    
                    # Redimensionar imagen
                    with PILImage.open(io.BytesIO(imagen_bytes)) as pil_img:
                        # Convertir a RGB si es necesario
                        if pil_img.mode == 'RGBA':
                            pil_img = pil_img.convert('RGB')
                        
                        # Redimensionar manteniendo proporción
                        pil_img.thumbnail((imagen_width, imagen_height), PILImage.Resampling.LANCZOS)
                        
                        # Guardar imagen redimensionada
                        output = io.BytesIO()
                        pil_img.save(output, format='JPEG', quality=85)
                        imagen_redimensionada = output.getvalue()
                    
                    # Crear objeto Image de openpyxl con configuración robusta para LibreOffice
                    img = ExcelImage(io.BytesIO(imagen_redimensionada))
                    
                    # Calcular posición (distribución en grid)
                    fila_offset = (i // max_cols) * 4  # 4 filas por cada fila de imágenes
                    col_offset = (i % max_cols) * 2    # 2 columnas por imagen
                    
                    fila_final = area_inicio + fila_offset
                    col_final = col_inicio + col_offset
                    
                    # Validar que no exceda el área calculada
                    if fila_final > area_fin:
                        logger.warning(f"⚠️ Imagen {i+1} excede área calculada, omitiendo")
                        continue
                    
                    # 🔧 CONFIGURACIÓN ESPECÍFICA PARA LIBREOFFICE
                    # Establecer dimensiones explícitas de la imagen (en puntos)
                    img.width = imagen_width * 0.75  # Convertir px a puntos (1px ≈ 0.75pt)
                    img.height = imagen_height * 0.75
                    
                    # Configurar celda de anclaje
                    celda = get_column_letter(col_final) + str(fila_final)
                    img.anchor = celda
                    
                    # 📏 AJUSTAR DIMENSIONES DE CELDAS PARA EVITAR SOLAPAMIENTO
                    self._configurar_celdas_para_imagen(ws, fila_final, col_final, imagen_width, imagen_height)
                    
                    # Insertar imagen
                    ws.add_image(img)
                    logger.info(f"✅ Imagen {i+1} insertada en {celda} con dimensiones {img.width:.1f}x{img.height:.1f}pt")
                    
                except Exception as e:
                    logger.error(f"❌ Error insertando imagen {i+1}: {str(e)}")
                    continue
            
            resultado = {
                'expandido': nueva_fila_firmas_calculada > 43,
                'nueva_fila_firmas': nueva_fila_firmas_calculada,
                'area_utilizada': (area_inicio, area_fin),
                'filas_totales': filas_necesarias,
                'imagenes_insertadas': min(len(imagenes), 6)
            }
            
            logger.info(f"🛡️ INSERCIÓN SEGURA COMPLETADA:")
            logger.info(f"   📊 Área utilizada: filas {area_inicio}-{area_fin}")
            logger.info(f"   🎯 Nueva posición firmas: fila {nueva_fila_firmas_calculada}")
            logger.info(f"   🖼️ Imágenes insertadas: {resultado['imagenes_insertadas']}")
            
            return resultado
            
        except Exception as e:
            logger.error(f"❌ Error crítico en inserción segura de imágenes: {str(e)}")
            # Retornar configuración por defecto en caso de error
            return {'expandido': False, 'nueva_fila_firmas': 43, 'area_utilizada': (40, 42)}

    def _configurar_celdas_para_imagen(self, ws, fila: int, columna: int, ancho_px: int, alto_px: int):
        """
        🔧 Configurar dimensiones de celdas específicamente para LibreOffice
        Evita superposición de imágenes al establecer alturas y anchos explícitos
        """
        try:
            # Convertir píxeles a unidades Excel
            # 1 píxel ≈ 0.75 puntos, Excel usa diferentes unidades
            ancho_excel = ancho_px * 0.14  # Aproximación para ancho de columna Excel
            alto_excel = alto_px * 0.75    # Puntos para altura de fila
            
            # 📏 CONFIGURAR ALTURA DE FILA (mínimo para que quepa la imagen)
            altura_minima = max(alto_excel, 30)  # Mínimo 30 puntos
            if ws.row_dimensions[fila].height is None or ws.row_dimensions[fila].height < altura_minima:
                ws.row_dimensions[fila].height = altura_minima
                logger.info(f"🔧 Altura fila {fila}: {altura_minima:.1f}pt")
            
            # 📐 CONFIGURAR ANCHO DE COLUMNA (solo si es necesario)
            col_letter = get_column_letter(columna)
            ancho_minimo = max(ancho_excel, 10)  # Mínimo 10 unidades Excel
            if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < ancho_minimo:
                ws.column_dimensions[col_letter].width = ancho_minimo
                logger.info(f"🔧 Ancho columna {col_letter}: {ancho_minimo:.1f}")
            
            # 🚀 CONFIGURACIÓN ADICIONAL PARA MÚLTIPLES FILAS (si la imagen es alta)
            filas_necesarias = max(1, int(alto_px / 100))  # Estimar filas necesarias
            for i in range(1, filas_necesarias):
                fila_extra = fila + i
                if fila_extra <= ws.max_row + 10:  # Limite razonable
                    altura_extra = max(alto_excel / filas_necesarias, 20)
                    ws.row_dimensions[fila_extra].height = altura_extra
                    logger.debug(f"🔧 Altura fila extra {fila_extra}: {altura_extra:.1f}pt")
            
        except Exception as e:
            logger.error(f"❌ Error configurando celdas para imagen: {str(e)}")

    def _configurar_alturas_firmas_libreoffice(self, ws, fila_firmas: int):
        """
        ⚡ Configurar alturas específicas para áreas de firma en LibreOffice
        Evita espacio excesivo en celdas combinadas de firmas
        """
        try:
            # 📏 ALTURAS ESPECÍFICAS PARA ÁREAS DE FIRMA
            altura_firma = 50  # Altura específica para firmas en puntos
            altura_nombre = 25  # Altura para nombres debajo de firmas
            
            # Área de firmas: 2 filas principales (firmas + nombres)
            fila_firma_actual = fila_firmas
            fila_nombres = fila_firmas + 1
            
            # ✍️ CONFIGURAR ALTURA FILA DE FIRMAS (Técnico A:D, Cliente E:H)
            if ws.row_dimensions[fila_firma_actual].height is None or ws.row_dimensions[fila_firma_actual].height > altura_firma:
                ws.row_dimensions[fila_firma_actual].height = altura_firma
                logger.info(f"⚡ Altura fila firmas {fila_firma_actual}: {altura_firma}pt (Técnico A:D, Cliente E:H)")
            
            # 📝 CONFIGURAR ALTURA FILA DE NOMBRES
            if ws.row_dimensions[fila_nombres].height is None or ws.row_dimensions[fila_nombres].height > altura_nombre:
                ws.row_dimensions[fila_nombres].height = altura_nombre
                logger.info(f"⚡ Altura fila nombres {fila_nombres}: {altura_nombre}pt")
            
            # 🚫 AJUSTAR FILAS ADICIONALES SI EXISTEN (limitar crecimiento)
            for fila_extra in range(fila_firmas + 2, fila_firmas + 5):  # 3 filas adicionales máximo
                if fila_extra <= ws.max_row:
                    altura_actual = ws.row_dimensions[fila_extra].height
                    if altura_actual is None or altura_actual > 20:  # Limitar altura excesiva
                        ws.row_dimensions[fila_extra].height = 20
                        logger.debug(f"⚡ Altura fila extra {fila_extra}: 20pt (limitada)")
            
            # 📐 CONFIGURAR ANCHOS DE COLUMNAS PARA FIRMAS (A:D y E:H)
            ancho_firma = 15  # Ancho adecuado para firmas
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                if ws.column_dimensions[col].width is None or ws.column_dimensions[col].width < ancho_firma:
                    ws.column_dimensions[col].width = ancho_firma
                    logger.debug(f"⚡ Ancho columna {col}: {ancho_firma}")
            
            logger.info(f"✅ Configuración de alturas de firma completada para fila {fila_firmas}")
            
        except Exception as e:
            logger.error(f"❌ Error configurando alturas de firmas: {str(e)}")

    def _configurar_contacto_solicitante(self, ws, contacto: str) -> bool:
        """
        📱 Configurar celda G20 para contacto solicitante con adaptación automática
        Evita desbordamiento mediante ajuste de texto, altura dinámica y ancho apropiado
        """
        try:
            if not contacto:
                logger.info("⚠️ No hay contacto solicitante para configurar")
                return False
            
            # Usar celda G20 para contacto solicitante
            celda_coord = 'G20'
            cell = ws[celda_coord]
            
            # Verificar si la celda está en un rango combinado
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
            
            # Establecer el valor
            cell.value = str(contacto)
            
            # 📱 CONFIGURACIÓN ESPECÍFICA PARA CONTACTO (puede ser largo)
            cell.alignment = Alignment(
                wrap_text=True,        # Activar ajuste de texto
                vertical='center',     # Centrado vertical
                horizontal='left',     # Alineación izquierda para mejor legibilidad
                shrink_to_fit=True     # Reducir tamaño de fuente si es necesario
            )
            
            # 📏 AJUSTAR ALTURA DE FILA DINÁMICAMENTE SEGÚN LONGITUD
            longitud_texto = len(contacto)
            if longitud_texto > 50:
                # Contacto muy largo: aumentar altura significativamente
                altura_necesaria = 45
                logger.info(f"📱 Contacto largo ({longitud_texto} chars), altura: {altura_necesaria}pt")
            elif longitud_texto > 25:
                # Contacto mediano: altura moderada
                altura_necesaria = 30
                logger.info(f"📱 Contacto mediano ({longitud_texto} chars), altura: {altura_necesaria}pt")
            else:
                # Contacto corto: altura estándar
                altura_necesaria = 20
                logger.info(f"📱 Contacto corto ({longitud_texto} chars), altura: {altura_necesaria}pt")
            
            # Aplicar altura si es mayor que la actual
            if ws.row_dimensions[20].height is None or ws.row_dimensions[20].height < altura_necesaria:
                ws.row_dimensions[20].height = altura_necesaria
            
            # 📐 AJUSTAR ANCHO DE COLUMNA G SI ES NECESARIO
            ancho_minimo = max(15, min(longitud_texto * 0.8, 25))  # Entre 15 y 25 unidades
            if ws.column_dimensions['G'].width is None or ws.column_dimensions['G'].width < ancho_minimo:
                ws.column_dimensions['G'].width = ancho_minimo
                logger.info(f"📐 Ancho columna G ajustado: {ancho_minimo}")
            
            logger.info(f"✅ Contacto solicitante configurado en {celda_coord}")
            logger.info(f"📝 Contenido: '{contacto[:50]}{'...' if len(contacto) > 50 else ''}'")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error configurando contacto solicitante: {str(e)}")
            return False

    def _obtener_firmas_conformidad(self, ot_id: int) -> dict:
        """
        🖊️ Obtener datos de firmas de conformidad desde la base de datos
        
        Args:
            ot_id: ID de la orden de trabajo
            
        Returns:
            dict: Información de las firmas (nombres e imágenes)
        """
        try:
            if not self.db_session:
                logger.warning("⚠️ No hay sesión de BD disponible para obtener firmas")
                return {
                    'nombre_tecnico': 'Por asignar',
                    'nombre_cliente': 'Cliente',
                    'firma_tecnico': None,
                    'firma_cliente': None,
                    'fecha_firma': None
                }
            
            # Importar modelo FirmaConformidad
            from app.models import FirmaConformidad
            
            # Buscar la firma más reciente para esta OT
            logger.info(f"🔍 Ejecutando consulta: FirmaConformidad.ot_id == {ot_id}")
            firma = self.db_session.query(FirmaConformidad).filter(
                FirmaConformidad.ot_id == ot_id
            ).order_by(FirmaConformidad.fecha_creacion.desc()).first()
            
            if not firma:
                logger.info(f"⚠️ No se encontraron firmas para OT {ot_id}")
                # También verificar todas las firmas disponibles para debug
                todas_firmas = self.db_session.query(FirmaConformidad).all()
                logger.info(f"📊 Total de firmas en BD: {len(todas_firmas)}")
                for f in todas_firmas[:5]:  # Solo las primeras 5
                    logger.info(f"   - Firma ID {f.numero_registro}: OT {f.ot_id}, Técnico: {f.nombre_tecnico}, Cliente: {f.nombre_cliente}")
                return {
                    'nombre_tecnico': 'Por asignar',
                    'nombre_cliente': 'Cliente',
                    'firma_tecnico': None,
                    'firma_cliente': None,
                    'fecha_firma': None
                }
            
            logger.info(f"✅ Firma encontrada: ID {firma.id}, Técnico: {firma.nombre_tecnico}, Cliente: {firma.nombre_cliente}")
            
            # Preparar datos de respuesta
            datos_firma = {
                'nombre_tecnico': firma.nombre_tecnico or 'Por asignar',
                'nombre_cliente': firma.nombre_cliente or 'Cliente',
                'fecha_firma': firma.fecha_firma
            }
            
            # 🎨 Procesar firmas de imagen (SOLO URLs de S3)
            if firma.firma_tecnico and firma.firma_tecnico != 'Sin firma':
                try:
                    if firma.firma_tecnico.startswith('https://') or firma.firma_tecnico.startswith('http://'):
                        # Es una URL de S3 - ÚNICO FORMATO SOPORTADO
                        datos_firma['firma_tecnico'] = firma.firma_tecnico
                        logger.info(f"📷 Firma técnico desde S3: {firma.firma_tecnico[:50]}...")
                    else:
                        # Solo aceptamos URLs de S3
                        datos_firma['firma_tecnico'] = None
                        logger.warning(f"⚠️ Firma técnico ignorada - solo se soportan URLs de S3")
                except Exception as e:
                    logger.warning(f"⚠️ Error procesando firma técnico: {e}")
                    datos_firma['firma_tecnico'] = None
            else:
                datos_firma['firma_tecnico'] = None

            if firma.firma_cliente and firma.firma_cliente != 'Sin firma':
                try:
                    if firma.firma_cliente.startswith('https://') or firma.firma_cliente.startswith('http://'):
                        # Es una URL de S3 - ÚNICO FORMATO SOPORTADO
                        datos_firma['firma_cliente'] = firma.firma_cliente
                        logger.info(f"📷 Firma cliente desde S3: {firma.firma_cliente[:50]}...")
                    else:
                        # Solo aceptamos URLs de S3
                        datos_firma['firma_cliente'] = None
                        logger.warning(f"⚠️ Firma cliente ignorada - solo se soportan URLs de S3")
                except Exception as e:
                    logger.warning(f"⚠️ Error procesando firma cliente: {e}")
                    datos_firma['firma_cliente'] = None
            else:
                datos_firma['firma_cliente'] = None            
            
            return datos_firma
            
        except Exception as e:
            logger.error(f"❌ Error obteniendo firmas de conformidad: {str(e)}")
            return {
                'nombre_tecnico': 'Por asignar',
                'nombre_cliente': 'Cliente',
                'firma_tecnico': None,
                'firma_cliente': None,
                'fecha_firma': None
            }

    def _insertar_imagenes_firmas(self, ws, firmas_data: dict, fila_firmas: int):
        """
        🎨 Insertar imágenes de firma en el Excel
        
        Args:
            ws: Worksheet de openpyxl
            firmas_data: Datos de las firmas obtenidos de la BD
            fila_firmas: Fila donde están los nombres de las firmas
        """
        try:
            logger.info(f"🎨 Insertando imágenes de firma en fila {fila_firmas}")
            
            # 🎯 USAR POSICIONES FIJAS ESPECÍFICAS: A48 para técnico, E48 para cliente (CENTRADAS)
            fila_imagenes_firmas = 48
            
            # Insertar firma del técnico en A48 (centrada) - SOLO S3
            if firmas_data.get('firma_tecnico'):
                self._insertar_imagen_firma_individual(
                    ws, 
                    firmas_data['firma_tecnico'], 
                    f'A{fila_imagenes_firmas}',
                    f'A{fila_imagenes_firmas}',
                    'técnico'
                )
            
            # Insertar firma del cliente en E48 (centrada) - SOLO S3
            if firmas_data.get('firma_cliente'):
                self._insertar_imagen_firma_individual(
                    ws, 
                    firmas_data['firma_cliente'], 
                    f'E{fila_imagenes_firmas}',
                    f'E{fila_imagenes_firmas}',
                    'cliente'
                )
            
            # Ajustar altura de la fila de imágenes de firma y anchos de columnas
            if firmas_data.get('firma_tecnico') or firmas_data.get('firma_cliente'):
                ws.row_dimensions[48].height = 80  # Altura para las firmas en fila 48
                logger.info(f"✅ Altura ajustada para fila de firmas 48: 80pt")
                
                # 🎯 CONFIGURAR ANCHOS DE COLUMNAS PARA CENTRADO PERFECTO
                # Columnas A-D para firma del técnico (expandir el área)
                for col in ['A', 'B', 'C', 'D']:
                    ws.column_dimensions[col].width = 25  # Ancho amplio para centrado
                
                # Columnas E-H para firma del cliente (expandir el área)  
                for col in ['E', 'F', 'G', 'H']:
                    ws.column_dimensions[col].width = 25  # Ancho amplio para centrado
                
                logger.info(f"✅ Anchos de columnas configurados para centrado de firmas: A-D y E-H = 25")
                
        except Exception as e:
            logger.error(f"❌ Error insertando imágenes de firma: {str(e)}")

    def _insertar_imagen_firma_individual(self, ws, firma_url: str, celda_inicio: str, celda_fin: str, tipo_firma: str):
        """
        🖊️ Insertar una imagen de firma individual desde URL de S3
        
        Args:
            ws: Worksheet de openpyxl
            firma_url: URL de S3 donde está almacenada la imagen
            celda_inicio: Celda superior izquierda (ej: 'A48')
            celda_fin: Celda inferior derecha (ej: 'A48')
            tipo_firma: 'técnico' o 'cliente' para logging
        """
        try:
            if not firma_url or not (firma_url.startswith('https://') or firma_url.startswith('http://')):
                logger.info(f"⚠️ URL de S3 inválida para firma de {tipo_firma}: {firma_url}")
                return
            
            # 🌐 Descargar imagen desde S3
            import requests
            response = requests.get(firma_url, timeout=10)
            if response.status_code != 200:
                logger.error(f"❌ Error descargando imagen de S3 para {tipo_firma}: HTTP {response.status_code}")
                return
            
            image_data = response.content
            logger.info(f"✅ Imagen descargada desde S3 para {tipo_firma}: {len(image_data)} bytes")
            
            # Crear imagen PIL para procesamiento
            pil_image = PILImage.open(io.BytesIO(image_data))
            
            # Redimensionar manteniendo aspecto (tamaño apropiado para firma)
            max_width, max_height = 200, 60  # Tamaño apropiado para firmas
            pil_image.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
            
            # Convertir a RGB si es necesario
            if pil_image.mode in ('RGBA', 'LA', 'P'):
                background = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                if pil_image.mode == 'P':
                    pil_image = pil_image.convert('RGBA')
                background.paste(pil_image, mask=pil_image.split()[-1] if 'A' in pil_image.mode else None)
                pil_image = background
            
            # Guardar en BytesIO para openpyxl
            img_buffer = io.BytesIO()
            pil_image.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Crear objeto imagen de openpyxl
            excel_image = ExcelImage(img_buffer)
            
            # Configurar posición y tamaño
            excel_image.anchor = celda_inicio
            excel_image.width = pil_image.width
            excel_image.height = pil_image.height
            
            # 🎯 CONFIGURAR ALINEAMIENTO CENTRADO DE LA CELDA Y LA IMAGEN
            from openpyxl.styles import Alignment
            cell = ws[celda_inicio]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 🎯 AJUSTAR ANCHO DE COLUMNA PARA MEJOR CENTRADO VISUAL
            col_letter = celda_inicio[0]  # A o E
            ws.column_dimensions[col_letter].width = 35  # Ancho mayor para mejor centrado
            
            # 🎯 CONFIGURAR POSICIONAMIENTO CENTRADO DE LA IMAGEN
            # Modificar las propiedades del anchor para centrado más preciso
            try:
                # Calcular offsets para centrado dentro de la celda
                cell_width_pts = ws.column_dimensions[col_letter].width * 7  # Aproximado en puntos
                cell_height_pts = ws.row_dimensions[48].height if ws.row_dimensions[48].height else 80
                
                # Offset horizontal para centrar (en puntos)
                h_offset = max(0, (cell_width_pts - excel_image.width) / 2)
                # Offset vertical para centrar (en puntos) 
                v_offset = max(0, (cell_height_pts - excel_image.height) / 2)
                
                # Crear anclaje centrado usando string con offset
                anchor_with_offset = f"{celda_inicio}"
                excel_image.anchor = anchor_with_offset
                
                # Configurar propiedades adicionales para centrado en LibreOffice/Excel
                if hasattr(excel_image, '_anchor'):
                    excel_image._anchor.col_off = int(h_offset * 12700)  # Convertir a EMU
                    excel_image._anchor.row_off = int(v_offset * 12700)  # Convertir a EMU
                
                logger.info(f"🎯 Configurado centrado con offsets: H={h_offset:.1f}pt, V={v_offset:.1f}pt")
                
            except Exception as offset_error:
                logger.warning(f"⚠️ No se pudo aplicar offset de centrado: {offset_error}")
                # Fallback: usar ancla simple
                excel_image.anchor = celda_inicio
            
            # Insertar en el worksheet
            ws.add_image(excel_image)
            
            logger.info(f"✅ Firma de {tipo_firma} insertada en {celda_inicio}:{celda_fin} ({pil_image.width}x{pil_image.height})")
            
        except Exception as e:
            logger.error(f"❌ Error insertando firma de {tipo_firma}: {str(e)}")

    def _insertar_imagenes_adjuntas_con_area_dinamica(self, ws, archivos_adjuntos: list) -> dict:
        """
        🎨 Insertar imágenes usando área dinámica expandida según cantidad y orientación
        
        Returns:
            dict: Información del área expandida utilizada
        """
        try:
            if not archivos_adjuntos or not self.s3_service:
                logger.info("⚠️ No hay archivos adjuntos para insertar o S3Service no disponible")
                return {'expandido': False, 'nueva_fila_firmas': 43}
            
            # Filtrar solo archivos de imagen de los adjuntos y limitar a máximo 3
            imagenes = []
            for archivo in archivos_adjuntos:
                nombre = archivo.get('nombre', '')
                if nombre and any(ext in nombre.lower() for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']):
                    imagenes.append(archivo)
                    # 🔢 LIMITAR A MÁXIMO 3 IMÁGENES PARA PRESERVAR ESTILO PDF
                    if len(imagenes) >= 3:
                        logger.info(f"📊 Limitando a 3 imágenes máximo para mantener estilo PDF consistente (área dinámica)")
                        break
            
            if not imagenes:
                logger.info("⚠️ No se encontraron archivos adjuntos de imagen para insertar")
                return {'expandido': False, 'nueva_fila_firmas': 43}
            
            logger.info(f"📷 Procesando {len(imagenes)} imagen(es) con área dinámica")
            
            # 🔍 ANÁLISIS DE ORIENTACIÓN PARA CÁLCULO DE ÁREA
            orientaciones = self._analizar_orientaciones_imagenes(imagenes)
            logger.info(f"📊 Orientaciones detectadas: {orientaciones}")
            
            # 🏗️ CALCULAR ÁREA DINÁMICA ÓPTIMA
            area_dinamica = self._calcular_area_dinamica_optimizada(len(imagenes), orientaciones)
            logger.info(f"🏗️ Área calculada: {area_dinamica['filas_totales']} filas ({area_dinamica['area_inicio_row']}-{area_dinamica['area_fin_row']})")
            if area_dinamica['expandido']:
                logger.info(f"📈 Expansión: {area_dinamica['razon_expansion']}")
            
            # 🎛️ CALCULAR LAYOUT CON ÁREA EXPANDIDA
            layout_config = self._calcular_layout_adaptativo_con_area(len(imagenes), orientaciones, area_dinamica)
            
            logger.info(f"📐 Layout final calculado:")
            logger.info(f"   📏 Dimensiones imagen: {layout_config['image_width']}x{layout_config['image_height']} px")
            logger.info(f"   📊 Distribución: {layout_config['cols_per_row']} columnas x {layout_config['max_rows']} filas")
            logger.info(f"   🎯 Tipo: {layout_config['layout_type']}")
            
            # 🖼️ INSERTAR IMÁGENES EN ÁREA EXPANDIDA
            for i, archivo in enumerate(imagenes[:layout_config['max_imagenes']]):
                try:
                    nombre = archivo.get('nombre', '')
                    logger.info(f"🎨 Procesando imagen {i+1}/{len(imagenes)}: {nombre}")
                    
                    # Descargar imagen
                    url_s3 = archivo.get('url', '')
                    imagen_bytes = self._descargar_imagen_desde_url_completa(url_s3)
                    if not imagen_bytes:
                        continue
                    
                    # Redimensionar con área expandida
                    imagen_redimensionada = self._redimensionar_imagen_optimizada(
                        imagen_bytes, 
                        layout_config['image_width'], 
                        layout_config['image_height'],
                        layout_config
                    )
                    if not imagen_redimensionada:
                        continue
                    
                    # Crear objeto Image de openpyxl
                    img = ExcelImage(io.BytesIO(imagen_redimensionada))
                    
                    # Calcular posición en área expandida
                    position = self._calcular_posicion_imagen(i, layout_config)
                    target_row = position['row']
                    target_col = position['col']
                    
                    # Verificar límites del área expandida
                    if target_row > area_dinamica['area_fin_row']:
                        logger.warning(f"⚠️ Imagen {i+1} excede área expandida, saltando")
                        continue
                    
                    # Anclar imagen
                    cell_coordinate = ws.cell(row=target_row, column=target_col).coordinate
                    img.anchor = cell_coordinate
                    
                    # Ajustar dimensiones con área expandida
                    self._ajustar_dimensiones_celda_con_area_expandida(ws, target_row, target_col, layout_config, area_dinamica)
                    
                    # Agregar imagen al worksheet
                    ws.add_image(img)
                    
                    logger.info(f"✅ Imagen {i+1} insertada en {cell_coordinate} (área expandida)")
                    
                except Exception as e:
                    logger.error(f"❌ Error insertando imagen {archivo.get('nombre', 'Sin nombre')}: {str(e)}")
                    continue
            
            logger.info(f"✅ Inserción completada usando área expandida")
            return area_dinamica
            
        except Exception as e:
            logger.error(f"❌ Error general insertando imágenes con área dinámica: {str(e)}")
            return {'expandido': False, 'nueva_fila_firmas': 43}
    
    def _insertar_imagenes_adjuntas(self, ws, archivos_adjuntos: list):
        """
        Descargar e insertar imágenes de archivos adjuntos en el Excel
        Las imágenes se insertan en la sección de archivos adjuntos (área A40:H42)
        """
        try:
            if not archivos_adjuntos or not self.s3_service:
                logger.info("⚠️ No hay archivos adjuntos para insertar o S3Service no disponible")
                return
            
            # Filtrar solo archivos de imagen de los adjuntos y limitar a máximo 3
            imagenes = []
            for archivo in archivos_adjuntos:
                nombre = archivo.get('nombre', '')
                if nombre and any(ext in nombre.lower() for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']):
                    imagenes.append(archivo)
                    # 🔢 LIMITAR A MÁXIMO 3 IMÁGENES PARA PRESERVAR ESTILO PDF
                    if len(imagenes) >= 3:
                        logger.info(f"📊 Limitando a 3 imágenes máximo para mantener estilo PDF consistente")
                        break
            
            if not imagenes:
                logger.info("⚠️ No se encontraron archivos adjuntos de imagen para insertar")
                return
            
            logger.info(f"📷 Insertando {len(imagenes)} imagen(es) en el Excel con distribución adaptativa")
            
            # 🔍 ANÁLISIS PREVIO DE ORIENTACIÓN DE IMÁGENES
            orientaciones = self._analizar_orientaciones_imagenes(imagenes)
            logger.info(f"📊 Orientaciones detectadas: {orientaciones}")
            
            # 🎨 SISTEMA DE DISTRIBUCIÓN ADAPTATIVO DE IMÁGENES
            # Área disponible: A40:H42 (8 columnas, 3 filas para archivos adjuntos)
            layout_config = self._calcular_layout_adaptativo(len(imagenes), orientaciones)
            
            logger.info(f"📐 Layout calculado para {len(imagenes)} imágenes:")
            logger.info(f"   📏 Dimensiones: {layout_config['image_width']}x{layout_config['image_height']} px")
            logger.info(f"   📊 Distribución: {layout_config['cols_per_row']} columnas x {layout_config['max_rows']} filas")
            logger.info(f"   🎯 Máximo imágenes: {layout_config['max_imagenes']}")
            
            for i, archivo in enumerate(imagenes[:layout_config['max_imagenes']]):
                try:
                    nombre = archivo.get('nombre', '')
                    logger.info(f"� Procesando archivo adjunto imagen {i+1}/{len(imagenes)}: {nombre}")
                    
                    # Obtener URL de S3 del archivo
                    url_s3 = archivo.get('url', '')
                    logger.info(f"🔗 URL S3: {url_s3}")
                    
                    # ESTRATEGIA PRINCIPAL: Descargar usando URL completa
                    imagen_bytes = self._descargar_imagen_desde_url_completa(url_s3)
                    if not imagen_bytes:
                        # FALLBACK: Buscar por nombre en carpetas
                        imagen_bytes = self._descargar_imagen_s3_con_fallback(nombre)
                    if not imagen_bytes:
                        logger.warning(f"⚠️ No se pudo descargar el archivo adjunto {nombre}")
                        continue
                    
                    # Redimensionar imagen según layout adaptativo optimizado
                    imagen_redimensionada = self._redimensionar_imagen_optimizada(
                        imagen_bytes, 
                        layout_config['image_width'], 
                        layout_config['image_height'],
                        layout_config  # Pasar configuración completa para optimización
                    )
                    if not imagen_redimensionada:
                        continue
                    
                    # Crear objeto Image de openpyxl
                    img = ExcelImage(io.BytesIO(imagen_redimensionada))
                    
                    # 🎯 POSICIONAMIENTO INTELIGENTE SEGÚN LAYOUT ADAPTATIVO
                    position = self._calcular_posicion_imagen(i, layout_config)
                    target_row = position['row']
                    target_col = position['col']
                        
                        # Verificar que no se sobreponga a las firmas
                    if target_row >= 46:
                            logger.warning(f"⚠️ Imagen {i+1} se saltará para evitar sobreposición con firmas")
                            continue
                    
                    # Anclar imagen a la celda centrada
                    cell_coordinate = ws.cell(row=target_row, column=target_col).coordinate
                    img.anchor = cell_coordinate
                    
                    # Ajustar dinámicamente las dimensiones según layout
                    self._ajustar_dimensiones_celda(ws, target_row, target_col, layout_config)
                    
                    # Agregar imagen al worksheet
                    ws.add_image(img)
                    
                    logger.info(f"✅ Imagen {i+1} insertada en {cell_coordinate} (fila {target_row}, col {target_col})")
                    logger.info(f"📐 Dimensiones: {layout_config['image_width']}x{layout_config['image_height']} px")
                    logger.info(f"🎨 Layout: {layout_config['layout_type']}")
                    
                except Exception as e:
                    logger.error(f"❌ Error insertando archivo adjunto imagen {archivo.get('nombre', 'Sin nombre')}: {str(e)}")
                    continue
            
            logger.info(f"✅ Proceso de inserción de imágenes completado")
            
        except Exception as e:
            logger.error(f"❌ Error general insertando imágenes: {str(e)}")
    
    def _analizar_orientaciones_imagenes(self, imagenes: list) -> dict:
        """
        🔍 Analizar orientaciones de las imágenes para optimizar el layout
        """
        orientaciones = {"vertical": 0, "horizontal": 0, "cuadrada": 0, "total": len(imagenes)}
        
        for i, archivo in enumerate(imagenes[:8]):  # Analizar máximo 8 imágenes
            try:
                url_s3 = archivo.get('url', '')
                if not url_s3:
                    continue
                    
                # Descargar solo header de imagen para obtener dimensiones sin descargar completa
                imagen_bytes = self._descargar_imagen_desde_url_completa(url_s3)
                if not imagen_bytes:
                    continue
                
                # Obtener dimensiones con PIL
                with PILImage.open(io.BytesIO(imagen_bytes)) as img:
                    width, height = img.size
                    aspect_ratio = width / height
                    
                    if aspect_ratio < 0.7:
                        orientaciones["vertical"] += 1
                    elif aspect_ratio > 1.4:
                        orientaciones["horizontal"] += 1
                    else:
                        orientaciones["cuadrada"] += 1
                        
                    logger.debug(f"📐 Imagen {i+1}: {width}x{height} (ratio: {aspect_ratio:.2f})")
                    
            except Exception as e:
                logger.debug(f"⚠️ No se pudo analizar imagen {i+1}: {str(e)}")
                continue
        
        # Calcular orientación predominante
        if orientaciones["vertical"] > orientaciones["horizontal"]:
            orientaciones["predominante"] = "vertical"
        elif orientaciones["horizontal"] > orientaciones["vertical"]:
            orientaciones["predominante"] = "horizontal" 
        else:
            orientaciones["predominante"] = "mixta"
            
        return orientaciones
    
    def _calcular_area_dinamica_optimizada(self, num_imagenes: int, orientaciones: dict = None) -> dict:
        """
        🏗️ Calcular área dinámica óptima para imágenes según cantidad y orientación
        
        Expande el área base A40:H42 (8 cols × 3 filas) cuando es necesario para mejorar estética
        
        Returns:
            dict: {
                'area_inicio_row': int,    # Fila de inicio del área
                'area_fin_row': int,       # Fila de fin del área  
                'area_inicio_col': int,    # Columna de inicio (siempre A=1)
                'area_fin_col': int,       # Columna de fin (siempre H=8)
                'filas_totales': int,      # Total de filas disponibles
                'columnas_totales': int,   # Total de columnas disponibles
                'nueva_fila_firmas': int,  # Nueva posición para firmas
                'expandido': bool,         # Si el área fue expandida
                'razon_expansion': str     # Razón de la expansión
            }
        """
        # Área base estándar: A40:H42 (3 filas × 8 columnas)
        area_base = {
            'area_inicio_row': 40,
            'area_fin_row': 42, 
            'area_inicio_col': 1,  # Columna A
            'area_fin_col': 8,     # Columna H
            'filas_totales': 5,    # 45-41+1 = 5 filas
            'columnas_totales': 8, # H-A+1 = 8 columnas
            'nueva_fila_firmas': 43,  # Posición original de firmas (después del área A40:H42)
            'expandido': False,
            'razon_expansion': 'Area base suficiente'
        }
        
        if num_imagenes <= 0:
            return area_base
        
        # Analizar orientación predominante
        es_predominante_vertical = orientaciones and orientaciones.get("predominante") == "vertical"
        porcentaje_vertical = orientaciones["vertical"] / orientaciones["total"] if orientaciones else 0
        
        # 🔍 LÓGICA DE EXPANSIÓN INTELIGENTE
        
        if num_imagenes == 1:
            # 1 imagen: usar área base, suficiente espacio
            return area_base
            
        elif num_imagenes == 2:
            if es_predominante_vertical:
                # 2 imágenes verticales apiladas: necesitan más altura
                return {
                    **area_base,
                    'area_fin_row': 47,        # +2 filas (41-47 = 7 filas)
                    'filas_totales': 7,
                    'nueva_fila_firmas': 48,
                    'expandido': True,
                    'razon_expansion': '2 imágenes verticales necesitan altura extra'
                }
            else:
                # 2 imágenes horizontales lado a lado: área base suficiente
                return area_base
                
        elif num_imagenes == 3:
            # 3 imágenes en distribución triangular: necesitan más espacio
            return {
                **area_base,
                'area_fin_row': 48,        # +3 filas (41-48 = 8 filas)
                'filas_totales': 8,
                'nueva_fila_firmas': 49,
                'expandido': True,
                'razon_expansion': '3 imágenes en layout triangular requieren espacio adicional'
            }
            
        elif num_imagenes == 4:
            if es_predominante_vertical:
                # 4 imágenes verticales en grid 2×2: necesitan altura considerable
                return {
                    **area_base,
                    'area_fin_row': 49,        # +4 filas (41-49 = 9 filas)
                    'filas_totales': 9,
                    'nueva_fila_firmas': 50,
                    'expandido': True,
                    'razon_expansion': '4 imágenes verticales en grid 2×2 necesitan altura extra'
                }
            else:
                # 4 imágenes horizontales: expansión moderada
                return {
                    **area_base,
                    'area_fin_row': 44,        # +2 filas (40-44 = 5 filas)
                    'filas_totales': 5,
                    'nueva_fila_firmas': 45,
                    'expandido': True,
                    'razon_expansion': '4 imágenes horizontales requieren espacio adicional'
                }
                
        elif num_imagenes <= 6:
            # 5-6 imágenes en grid 3×2: necesitan expansión significativa
            filas_extra = 4 if es_predominante_vertical else 3
            return {
                **area_base,
                'area_fin_row': 42 + filas_extra,
                'filas_totales': 3 + filas_extra,
                'nueva_fila_firmas': 43 + filas_extra,
                'expandido': True,
                'razon_expansion': f'{num_imagenes} imágenes en grid 3×2 requieren {filas_extra} filas extra'
            }
            
        else:
            # 7+ imágenes: expansión máxima para grid compacto
            filas_extra = 5 if es_predominante_vertical else 4
            return {
                **area_base,
                'area_fin_row': 42 + filas_extra,
                'filas_totales': 3 + filas_extra,
                'nueva_fila_firmas': 43 + filas_extra,
                'expandido': True,
                'razon_expansion': f'{num_imagenes} imágenes en grid compacto requieren {filas_extra} filas extra'
            }
    
    def _calcular_layout_adaptativo(self, num_imagenes: int, orientaciones: dict = None) -> dict:
        """
        🎨 Calcular distribución adaptativa de imágenes según cantidad y orientación
        
        Área disponible: A40:H42 (8 columnas × 3 filas = 24 celdas)
        
        Distribuciones optimizadas para área más compacta:
        - 1 imagen: Centrada (4x2 celdas)
        - 2 imágenes: Lado a lado horizontalmente (4x3 cada una)
        - 3+ imágenes: Grid horizontal compacto
        """
        # Área base disponible (celdas de Excel)
        area_inicio_col = 1  # Columna A = 1
        area_fin_col = 8     # Columna H = 8
        area_inicio_row = 40 # Fila 40
        area_fin_row = 42    # Fila 42
        
        area_ancho_celdas = area_fin_col - area_inicio_col + 1  # 8 columnas
        area_alto_celdas = area_fin_row - area_inicio_row + 1   # 3 filas
        
        # Detectar si hay orientación predominante
        es_predominante_vertical = orientaciones and orientaciones.get("predominante") == "vertical"
        porcentaje_vertical = orientaciones["vertical"] / orientaciones["total"] if orientaciones else 0
        
        if num_imagenes <= 0:
            return {"max_imagenes": 0}
        
        elif num_imagenes == 1:
            # 1 imagen: Centrada, optimizada según orientación
            if es_predominante_vertical:
                # Para imagen vertical: más altura, menos anchura, posición optimizada
                return {
                    "max_imagenes": 1,
                    "cols_per_row": 1,
                    "max_rows": 1,
                    "image_width": 180,   # Más estrecho para imagen alta
                    "image_height": 160,  # Altura ajustada para área A40:H42 (3 filas)
                    "layout_type": "single_vertical",
                    "spacing": {"row": 0, "col": 0},
                    "start_row": 40,      # Desde arriba del área A40:H42
                    "start_col": 4        # Centrada horizontalmente (columna D)
                }
            else:
                # Para imagen horizontal o cuadrada: configuración balanceada
                return {
                    "max_imagenes": 1,
                    "cols_per_row": 1,
                    "max_rows": 1,
                    "image_width": 320,   # Más ancho para aprovechar espacio horizontal
                    "image_height": 140,  # Altura ajustada para área A40:H42
                    "layout_type": "single_horizontal",
                    "spacing": {"row": 0, "col": 0},
                    "start_row": 41,      # Centrada verticalmente (fila 41 de 40-42)
                    "start_col": 3        # Más centrada (columna C, dejando espacio A-B y F-H)
                }
            
        elif num_imagenes == 2:
            # 2 imágenes: Layout adaptativo según orientación predominante
            if es_predominante_vertical or porcentaje_vertical > 0.5:
                # Para imágenes verticales: layout apilado (una arriba, otra abajo)
                return {
                    "max_imagenes": 2,
                    "cols_per_row": 1,  # Una por fila para imágenes altas
                    "max_rows": 2,
                    "image_width": 220,   # Más ancho disponible
                    "image_height": 100,  # Altura controlada para área compacta A40:H42
                    "layout_type": "vertical_stack",
                    "spacing": {"row": 1, "col": 0},
                    "start_row": 40,      # Primera en fila 40
                    "start_col": 4        # Centrada horizontalmente
                }
            else:
                # Para imágenes horizontales: layout lado a lado
                return {
                    "max_imagenes": 2,
                    "cols_per_row": 2, 
                    "max_rows": 1,
                    "image_width": 180,   # Ancho para que quepan lado a lado
                    "image_height": 140,  # Altura ajustada para área A40:H42
                    "layout_type": "horizontal_pair",
                    "spacing": {"row": 0, "col": 3},
                    "start_row": 41,      # Centrada verticalmente (fila 41 de 40-42)
                    "start_col": 2        # Empezar en columna B
                }
            
        elif num_imagenes == 3:
            # 3 imágenes: Layout adaptativo según orientación
            if es_predominante_vertical:
                # Para imágenes verticales: configuración más alta y compacta
                return {
                    "max_imagenes": 3,
                    "cols_per_row": 2,  # Máximo 2 por fila  
                    "max_rows": 2,
                    "image_width": 160,   # Más compacto horizontalmente
                    "image_height": 160,  # Más altura para verticales
                    "layout_type": "triangular_vertical",
                    "spacing": {"row": 1, "col": 2},
                    "start_row": 41,      # Desde arriba
                    "start_col": 2
                }
            else:
                # Para imágenes horizontales: configuración estándar
                return {
                    "max_imagenes": 3,
                    "cols_per_row": 2,  # Máximo 2 por fila
                    "max_rows": 2,
                    "image_width": 180,
                    "image_height": 140,
                    "layout_type": "triangular_horizontal",
                    "spacing": {"row": 2, "col": 2},
                    "start_row": 41,
                    "start_col": 2
                }
            
        elif num_imagenes == 4:
            # 4 imágenes: Grid 2×2 adaptativo según orientación
            if es_predominante_vertical:
                # Para imágenes verticales: grid más alto y compacto
                return {
                    "max_imagenes": 4,
                    "cols_per_row": 2,
                    "max_rows": 2,
                    "image_width": 140,   # Más compacto para verticales
                    "image_height": 140,  # Más cuadrado para balance
                    "layout_type": "grid_2x2_vertical",
                    "spacing": {"row": 1, "col": 2},
                    "start_row": 41,
                    "start_col": 2
                }
            else:
                # Para imágenes horizontales: grid estándar
                return {
                    "max_imagenes": 4,
                    "cols_per_row": 2,
                    "max_rows": 2,
                    "image_width": 160,
                    "image_height": 120,
                    "layout_type": "grid_2x2_horizontal",
                    "spacing": {"row": 2, "col": 3},
                    "start_row": 41,
                    "start_col": 2
                }
            
        elif num_imagenes <= 6:
            # 5-6 imágenes: Grid 3×2 adaptativo según orientación
            if es_predominante_vertical:
                # Para imágenes verticales: más altura, menos anchura
                return {
                    "max_imagenes": 6,
                    "cols_per_row": 3,
                    "max_rows": 2,
                    "image_width": 120,   # Más compacto para verticales
                    "image_height": 120,  # Más cuadrado para balance
                    "layout_type": "grid_3x2_vertical",
                    "spacing": {"row": 1, "col": 1},
                    "start_row": 41,
                    "start_col": 1
                }
            else:
                # Para imágenes horizontales: configuración estándar
                return {
                    "max_imagenes": 6,
                    "cols_per_row": 3,
                    "max_rows": 2,
                    "image_width": 140,
                    "image_height": 100,
                    "layout_type": "grid_3x2_horizontal",
                    "spacing": {"row": 2, "col": 2},
                    "start_row": 41,
                    "start_col": 1
                }
            
        else:
            # 7+ imágenes: Grid compacto adaptativo
            if es_predominante_vertical:
                # Para imágenes verticales: grid más compacto y alto
                return {
                    "max_imagenes": 8,
                    "cols_per_row": 4,
                    "max_rows": 2,
                    "image_width": 100,   # Muy compacto para verticales
                    "image_height": 100,  # Más cuadrado
                    "layout_type": "grid_compact_vertical",
                    "spacing": {"row": 1, "col": 1},
                    "start_row": 41,
                    "start_col": 1
                }
            else:
                # Para imágenes horizontales: grid estándar compacto
                return {
                    "max_imagenes": 8,
                    "cols_per_row": 4,
                    "max_rows": 2,
                    "image_width": 120,
                    "image_height": 80,
                    "layout_type": "grid_compact_horizontal",
                    "spacing": {"row": 2, "col": 1},
                    "start_row": 41,
                    "start_col": 1
                }
    
    def _calcular_posicion_imagen(self, indice: int, layout_config: dict) -> dict:
        """
        🎯 Calcular posición específica de una imagen según el layout adaptativo
        """
        layout_type = layout_config["layout_type"]
        
        # Layouts de imagen única
        if layout_type in ["single_vertical", "single_horizontal"]:
            return {
                "row": layout_config["start_row"],
                "col": layout_config["start_col"]
            }
        
        # Layout apilado vertical (para 2 imágenes verticales)
        elif layout_type == "vertical_stack":
            return {
                "row": layout_config["start_row"] + (indice * (layout_config["spacing"]["row"] + 2)),
                "col": layout_config["start_col"]
            }
        
        # Layouts tradicionales (grid-based)
        elif layout_type in ["horizontal_pair", "triangular_vertical", "triangular_horizontal", 
                           "grid_2x2_vertical", "grid_2x2_horizontal", "grid_3x2_vertical", 
                           "grid_3x2_horizontal", "grid_compact_vertical", "grid_compact_horizontal"]:
            # Cálculo estándar de posición en grid
            fila_actual = indice // layout_config["cols_per_row"]
            col_en_fila = indice % layout_config["cols_per_row"]
            
            # Para layout triangular: ajustar primera fila para centrar imagen única
            if "triangular" in layout_type and fila_actual == 0:
                # Primera fila centrada (solo 1 imagen)
                col_offset = 1  # Desplazar una columna a la derecha para centrar
            else:
                col_offset = 0
            
            return {
                "row": layout_config["start_row"] + (fila_actual * (layout_config["spacing"]["row"] + 1)),
                "col": layout_config["start_col"] + (col_en_fila * layout_config["spacing"]["col"]) + col_offset
            }
        
        # Fallback para tipos no reconocidos
        else:
            return {
                "row": layout_config["start_row"],
                "col": layout_config["start_col"]
            }
    
    def _ajustar_dimensiones_celda_mejorada(self, ws, target_row: int, target_col: int, layout_config: dict):
        """
        ⚖️ Ajustar dinámicamente las dimensiones de filas y columnas para imágenes con cálculo preciso
        Área disponible: A40:H42 (8 columnas × 3 filas)
        """
        try:
            image_width = layout_config["image_width"]
            image_height = layout_config["image_height"]
            layout_type = layout_config.get("layout_type", "unknown")
            
            logger.info(f"🔧 Ajustando dimensiones para imagen {image_width}x{image_height}px en {layout_type}")
            
            # 📐 ÁREA DISPONIBLE Y LÍMITES
            area_max_row = 42  # Hasta fila 42 (área de archivos adjuntos)
            area_max_col = 8   # Hasta columna H
            filas_disponibles = area_max_row - target_row + 1
            columnas_disponibles = area_max_col - target_col + 1
            
            # 📏 CÁLCULO MEJORADO DE ALTURA DE FILAS
            # Conversión más precisa: 1 píxel ≈ 0.75 puntos, altura Excel en puntos
            if layout_type == "single_horizontal":
                # Para imagen horizontal única: usar 2-3 filas con altura generosa
                filas_necesarias = min(3, filas_disponibles)
                altura_por_fila = max(35, image_height * 0.8 / filas_necesarias)  
            elif layout_type == "single_vertical":
                # Para imagen vertical única: usar máximo 4 filas con altura amplia
                filas_necesarias = min(4, filas_disponibles)
                altura_por_fila = max(40, image_height * 0.9 / filas_necesarias)
            else:
                # Para múltiples imágenes: distribuir altura disponible
                filas_necesarias = min(layout_config.get("max_rows", 2), filas_disponibles)
                altura_por_fila = max(25, image_height * 0.7 / filas_necesarias)
            
            # 📐 CÁLCULO MEJORADO DE ANCHO DE COLUMNAS  
            # Conversión: 1 píxel ≈ 0.14 unidades de columna Excel
            if layout_type == "single_horizontal":
                # Para imagen horizontal única: usar 3-4 columnas con ancho generoso
                columnas_necesarias = min(4, columnas_disponibles)
                ancho_por_columna = max(20, image_width * 0.12 / columnas_necesarias)
            elif layout_type == "single_vertical":
                # Para imagen vertical única: usar 2-3 columnas más estrechas
                columnas_necesarias = min(3, columnas_disponibles)
                ancho_por_columna = max(18, image_width * 0.15 / columnas_necesarias)
            else:
                # Para múltiples imágenes: distribuir ancho disponible
                columnas_necesarias = min(layout_config.get("cols_per_row", 2), columnas_disponibles)
                ancho_por_columna = max(15, image_width * 0.1 / columnas_necesarias)
            
            logger.info(f"📊 Distribución calculada: {filas_necesarias}filas×{altura_por_fila:.1f}h, {columnas_necesarias}cols×{ancho_por_columna:.1f}w")
            
            # 🔧 APLICAR AJUSTES DE FILAS
            for i in range(filas_necesarias):
                row_num = target_row + i
                if row_num <= area_max_row:
                    # Asegurar altura mínima para visibilidad
                    altura_final = max(altura_por_fila, 30)
                    ws.row_dimensions[row_num].height = altura_final
                    logger.debug(f"   Fila {row_num}: altura {altura_final:.1f}")
                else:
                    logger.warning(f"⚠️ Fila {row_num} excede límite del área")
                    break
            
            # 🔧 APLICAR AJUSTES DE COLUMNAS
            from openpyxl.utils import get_column_letter
            for i in range(columnas_necesarias):
                col_num = target_col + i
                if col_num <= area_max_col:
                    col_letter = get_column_letter(col_num)
                    # Asegurar ancho mínimo para visibilidad
                    ancho_final = max(ancho_por_columna, 15)
                    ws.column_dimensions[col_letter].width = ancho_final
                    logger.debug(f"   Columna {col_letter}: ancho {ancho_final:.1f}")
                else:
                    logger.warning(f"⚠️ Columna {col_num} excede límite del área")
                    break
            
            logger.info(f"✅ Dimensiones aplicadas: {filas_necesarias} filas desde {target_row}, {columnas_necesarias} columnas desde {get_column_letter(target_col)}")
            
        except Exception as e:
            logger.error(f"❌ Error ajustando dimensiones de celda: {str(e)}")
    
    def _ajustar_dimensiones_celda(self, ws, target_row: int, target_col: int, layout_config: dict):
        """
        Función legacy - usar _ajustar_dimensiones_celda_mejorada para mejor rendimiento
        """
        return self._ajustar_dimensiones_celda_mejorada(ws, target_row, target_col, layout_config)
    
    def _calcular_layout_adaptativo_con_area(self, num_imagenes: int, orientaciones: dict, area_dinamica: dict) -> dict:
        """
        🎛️ Calcular layout adaptativo usando el área dinámica expandida
        
        Similar a _calcular_layout_adaptativo pero usa el área expandida como referencia
        """
        # Obtener layout base
        layout_base = self._calcular_layout_adaptativo(num_imagenes, orientaciones)
        
        # Actualizar con área expandida
        layout_expandido = {
            **layout_base,
            'start_row': area_dinamica['area_inicio_row'],  # Usar inicio del área expandida
            'area_max_row': area_dinamica['area_fin_row'],  # Límite máximo
            'area_filas_disponibles': area_dinamica['filas_totales'],  # Total de filas disponibles
            'area_expandida': area_dinamica['expandido'],   # Indicador de expansión
        }
        
        # 🔧 AJUSTES ESPECÍFICOS PARA ÁREA EXPANDIDA
        if area_dinamica['expandido']:
            # Para área expandida, podemos usar imágenes más grandes
            factor_expansion = min(1.3, area_dinamica['filas_totales'] / 5)  # Máximo 30% más grande
            
            layout_expandido['image_width'] = int(layout_expandido['image_width'] * factor_expansion)
            layout_expandido['image_height'] = int(layout_expandido['image_height'] * factor_expansion)
            
            logger.info(f"🔧 Layout ajustado para área expandida (factor: {factor_expansion:.2f})")
            logger.info(f"   📏 Nuevas dimensiones: {layout_expandido['image_width']}x{layout_expandido['image_height']}px")
        
        return layout_expandido
    
    def _ajustar_dimensiones_celda_con_area_expandida(self, ws, target_row: int, target_col: int, layout_config: dict, area_dinamica: dict):
        """
        ⚖️ Ajustar dimensiones de celdas considerando el área expandida disponible
        """
        try:
            image_width = layout_config["image_width"]
            image_height = layout_config["image_height"]
            layout_type = layout_config.get("layout_type", "unknown")
            
            # Usar límites del área expandida
            area_max_row = area_dinamica['area_fin_row']
            area_max_col = area_dinamica['area_fin_col']
            filas_disponibles = area_max_row - target_row + 1
            columnas_disponibles = area_max_col - target_col + 1
            
            logger.info(f"🔧 Ajustando con área expandida: filas hasta {area_max_row}, {filas_disponibles} disponibles")
            
            # 📏 CÁLCULO OPTIMIZADO PARA ÁREA EXPANDIDA
            if area_dinamica['expandido']:
                # Con área expandida, podemos ser más generosos
                if layout_type == "single_horizontal":
                    filas_necesarias = min(4, filas_disponibles)  # +1 fila vs área normal
                    altura_por_fila = max(40, image_height * 0.9 / filas_necesarias)  # +5 vs área normal
                elif layout_type == "single_vertical":
                    filas_necesarias = min(5, filas_disponibles)  # +1 fila vs área normal
                    altura_por_fila = max(45, image_height * 0.95 / filas_necesarias)  # +5 vs área normal
                else:
                    filas_necesarias = min(layout_config.get("max_rows", 3), filas_disponibles)
                    altura_por_fila = max(30, image_height * 0.8 / filas_necesarias)  # +5 vs área normal
                
                # Ancho también más generoso
                if layout_type == "single_horizontal":
                    columnas_necesarias = min(5, columnas_disponibles)  # +1 columna
                    ancho_por_columna = max(25, image_width * 0.15 / columnas_necesarias)  # +5
                elif layout_type == "single_vertical":
                    columnas_necesarias = min(4, columnas_disponibles)  # +1 columna
                    ancho_por_columna = max(22, image_width * 0.18 / columnas_necesarias)  # +4
                else:
                    columnas_necesarias = min(layout_config.get("cols_per_row", 3), columnas_disponibles)
                    ancho_por_columna = max(18, image_width * 0.12 / columnas_necesarias)  # +3
            else:
                # Usar cálculo normal
                return self._ajustar_dimensiones_celda_mejorada(ws, target_row, target_col, layout_config)
            
            logger.info(f"📊 Área expandida: {filas_necesarias}filas×{altura_por_fila:.1f}h, {columnas_necesarias}cols×{ancho_por_columna:.1f}w")
            
            # 🔧 APLICAR AJUSTES EXPANDIDOS
            for i in range(filas_necesarias):
                row_num = target_row + i
                if row_num <= area_max_row:
                    altura_final = max(altura_por_fila, 35)  # Altura mínima mayor
                    ws.row_dimensions[row_num].height = altura_final
                else:
                    break
            
            from openpyxl.utils import get_column_letter
            for i in range(columnas_necesarias):
                col_num = target_col + i
                if col_num <= area_max_col:
                    col_letter = get_column_letter(col_num)
                    ancho_final = max(ancho_por_columna, 18)  # Ancho mínimo mayor
                    ws.column_dimensions[col_letter].width = ancho_final
                else:
                    break
            
            logger.info(f"✅ Dimensiones aplicadas con área expandida")
            
        except Exception as e:
            logger.error(f"❌ Error ajustando dimensiones con área expandida: {str(e)}")
            # Fallback a función normal
            return self._ajustar_dimensiones_celda_mejorada(ws, target_row, target_col, layout_config)
    
    def _insertar_imagen_original(self, ws, imagen_original: str):
        """
        Insertar la imagen original de la solicitud en el área específica B30:C31 (centrada)
        Esta es la imagen que se subió con la solicitud B2C/B2B original
        """
        try:
            if not imagen_original or not imagen_original.strip():
                logger.info("⚠️ No hay imagen original de la solicitud para insertar")
                return
            
            if not self.s3_service:
                logger.info("⚠️ S3Service no disponible para imagen original")
                return
            
            # Extraer el nombre del archivo de la URL
            nombre_imagen = imagen_original.split('/')[-1] if '/' in imagen_original else imagen_original
            
            # Verificar que sea una imagen
            if not any(ext in nombre_imagen.lower() for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']):
                logger.info(f"⚠️ El archivo {nombre_imagen} no es una imagen reconocida")
                return
            
            logger.info(f"🖼️ Procesando imagen original de la solicitud: {nombre_imagen}")
            
            logger.info(f"🔗 URL S3 imagen original: {imagen_original}")
            
            # Descargar imagen directamente usando la URL completa de S3
            imagen_bytes = self._descargar_imagen_desde_url_completa(imagen_original)
            if not imagen_bytes:
                logger.warning(f"⚠️ No se pudo descargar la imagen original: {nombre_imagen}")
                return
            
            # Redimensionar imagen para el área B30:C31 (más pequeña que los adjuntos)
            image_width = 200   # Ancho apropiado para el área de imagen original
            image_height = 120  # Alto que cabe en A30:A32 (3 filas)
            
            imagen_redimensionada = self._redimensionar_imagen(imagen_bytes, image_width, image_height)
            if not imagen_redimensionada:
                logger.warning(f"⚠️ No se pudo redimensionar la imagen original: {nombre_imagen}")
                return
            
            # Crear objeto Image de openpyxl
            img = ExcelImage(io.BytesIO(imagen_redimensionada))
            
            # Posición específica para la imagen original: B30:C31 (más centrada)
            target_row = 30  # Fila 30 (inicio del área de imagen original)
            target_col = 2   # Columna B (más centrada que A)
            
            # Anclar imagen a la celda B30
            cell_coordinate = ws.cell(row=target_row, column=target_col).coordinate
            img.anchor = cell_coordinate
            
            # Ajustar altura de las filas B30:C31 para acomodar la imagen
            for row_num in range(30, 32):  # Filas 30, 31
                ws.row_dimensions[row_num].height = 25  # Altura suficiente para la imagen
            
            # Agregar imagen al worksheet
            ws.add_image(img)
            
            logger.info(f"✅ Imagen original de solicitud {nombre_imagen} insertada centrada en {cell_coordinate} (área B30:C31)")
            logger.info(f"📐 Dimensiones imagen original: {image_width}x{image_height} px")
            
        except Exception as e:
            logger.error(f"❌ Error insertando imagen original de la solicitud: {str(e)}")
    
    def _descargar_imagen_s3(self, nombre_archivo: str) -> Optional[bytes]:
        """
        Descargar imagen de S3
        """
        try:
            if not self.s3_service:
                logger.error("S3Service no disponible")
                return None
            
            logger.info(f"🔄 Descargando imagen desde S3: {nombre_archivo}")
            
            # Debug: Intentar encontrar el archivo con variaciones del nombre
            posibles_nombres = [
                nombre_archivo,  # Nombre original
                f"images/{nombre_archivo}",  # En carpeta images
                f"uploads/{nombre_archivo}",  # En carpeta uploads
                f"b2c/{nombre_archivo}",  # En carpeta b2c
            ]
            
            for nombre_prueba in posibles_nombres:
                logger.info(f"🔍 Intentando descargar con nombre: {nombre_prueba}")
                
                # Método 1: Intentar descargar directamente
                download_result = self.s3_service.download_file_from_s3(nombre_prueba)
                if download_result.get('success', False):
                    content = download_result.get('content')
                    if content:
                        logger.info(f"✅ Imagen encontrada y descargada: {nombre_prueba} ({len(content)} bytes)")
                        return content
                else:
                    logger.debug(f"❌ No encontrado con nombre: {nombre_prueba} - {download_result.get('error', 'Sin error')}")
            
            # Si no se encuentra con ninguna variación, listar archivos para debug
            logger.warning(f"⚠️ Archivo no encontrado con ninguna variación. Listando archivos S3 para debug...")
            self._debug_listar_archivos_s3(nombre_archivo)
            
            # Método 2: Fallback usando URL presignada con nombre original
            logger.info(f"🔄 Intentando fallback con URL presignada: {nombre_archivo}")
            url_result = self.s3_service.generate_presigned_download_url(nombre_archivo)
            if url_result.get('success', False):
                url_descarga = url_result.get('url')
                if url_descarga:
                    response = requests.get(url_descarga, timeout=30)
                    response.raise_for_status()
                    logger.info(f"✅ Imagen descargada via URL presignada: {nombre_archivo} ({len(response.content)} bytes)")
                    return response.content
            
            logger.error(f"❌ No se pudo encontrar la imagen {nombre_archivo} en S3")
            return None
            
        except Exception as e:
            logger.error(f"❌ Error descargando imagen {nombre_archivo} de S3: {str(e)}")
            return None
    
    def _descargar_imagen_s3_con_fallback(self, nombre_archivo: str) -> Optional[bytes]:
        """
        Sistema de fallback escalable para buscar imágenes en las carpetas específicas de S3:
        - b2c/ : Solicitudes B2C (prefijo b2c_)
        - b2b/ : Solicitudes B2B (prefijo b2b_) 
        - ots/ : Archivos adjuntos OTs (prefijo ot_)
        - planta_san_pedro/ : Planta San Pedro (prefijo psp_)
        """
        try:
            logger.info(f"🔍 Iniciando búsqueda escalable para: {nombre_archivo}")
            
            # Estrategia 1: Intentar descarga directa (por si está en root)
            resultado = self._descargar_imagen_s3(nombre_archivo)
            if resultado:
                logger.info(f"✅ Imagen encontrada en root: {nombre_archivo}")
                return resultado
            
            # Estrategia 2: Buscar en carpetas específicas según estructura de S3
            carpetas_especificas = [
                'b2c/',           # Solicitudes B2C
                'b2b/',           # Solicitudes B2B  
                'ots/',           # Archivos adjuntos de OTs
                'planta_san_pedro/'  # Planta San Pedro
            ]
            
            for carpeta in carpetas_especificas:
                nombre_con_carpeta = f"{carpeta}{nombre_archivo}"
                resultado = self._descargar_imagen_s3(nombre_con_carpeta)
                if resultado:
                    logger.info(f"✅ Imagen encontrada en carpeta '{carpeta}': {nombre_con_carpeta}")
                    return resultado
            
            # Estrategia 3: Buscar archivos con prefijos específicos en cada carpeta
            # Esto cubre casos donde el archivo tiene prefijo completo
            patrones_busqueda = [
                # B2C: b2c_YYYYMMDD_HHMMSS_hash_nombre
                f"b2c/b2c_*{nombre_archivo}*",
                # B2B: b2b_YYYYMMDD_HHMMSS_hash_nombre  
                f"b2b/b2b_*{nombre_archivo}*",
                # OTs: ot_NUMERO_YYYYMMDD_HHMMSS_hash_nombre
                f"ots/ot_*{nombre_archivo}*",
                # Planta San Pedro: psp_NUMERO_YYYYMMDD_HHMMSS_hash_nombre
                f"planta_san_pedro/psp_*{nombre_archivo}*"
            ]
            
            # Para esta estrategia, intentar con variaciones del nombre
            nombres_alternativos = [
                nombre_archivo,
                nombre_archivo.replace(' ', '_'),
                nombre_archivo.replace(' ', '-'), 
                nombre_archivo.replace(' ', ''),
                nombre_archivo.replace(' ', '%20')
            ]
            
            for nombre_alt in nombres_alternativos:
                for carpeta in carpetas_especificas:
                    # Buscar archivos que contengan el nombre en cada carpeta
                    if self.s3_service:
                        try:
                            # Listar archivos en la carpeta específica
                            list_result = self.s3_service.list_files(folder=carpeta.rstrip('/'), limit=50)
                            if list_result.get('success', False):
                                files = list_result.get('files', [])
                                # Buscar archivos que contengan el nombre (sin extension)
                                nombre_base = nombre_alt.lower().replace('.png', '').replace('.jpg', '').replace('.jpeg', '')
                                
                                for file_info in files:
                                    file_key = file_info.get('key', '').lower()
                                    if nombre_base in file_key and any(ext in file_key for ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']):
                                        # Intentar descargar este archivo
                                        full_key = file_info.get('key', '')
                                        resultado = self._descargar_imagen_s3(full_key)
                                        if resultado:
                                            logger.info(f"✅ Imagen encontrada por búsqueda en {carpeta}: {full_key}")
                                            return resultado
                        except Exception as e:
                            logger.debug(f"Error buscando en carpeta {carpeta}: {str(e)}")
                            continue
            
            logger.warning(f"❌ No se pudo encontrar la imagen '{nombre_archivo}' en ninguna carpeta de S3")
            return None
            
        except Exception as e:
            logger.error(f"❌ Error en búsqueda escalable para {nombre_archivo}: {str(e)}")
            return None
    
    def _descargar_imagen_desde_url_completa(self, url_s3: str) -> Optional[bytes]:
        """
        Descargar imagen directamente desde URL completa de S3
        Sistema escalable que maneja todas las carpetas: b2c/, b2b/, ots/, planta_san_pedro/
        """
        try:
            if not url_s3 or not url_s3.strip():
                logger.warning("❌ URL de S3 vacía")
                return None
            
            logger.info(f"🔄 Descargando imagen desde URL completa: {url_s3}")
            
            # Estrategia 1: Intentar descarga directa con la URL proporcionada
            try:
                response = requests.get(url_s3, timeout=30)
                response.raise_for_status()
                
                if response.content:
                    logger.info(f"✅ Imagen descargada exitosamente desde URL directa ({len(response.content)} bytes)")
                    return response.content
                    
            except requests.exceptions.RequestException as e:
                logger.warning(f"⚠️ Descarga directa falló: {str(e)}")
                
                # Estrategia 2: Si la URL directa falla, intentar con S3Service
                if self.s3_service:
                    logger.info("🔄 Intentando descarga alternativa con S3Service...")
                    return self._descargar_con_s3service_como_fallback(url_s3)
            
            logger.warning("❌ Respuesta vacía desde URL de S3")
            return None
                
        except Exception as e:
            logger.error(f"❌ Error general descargando desde URL {url_s3}: {str(e)}")
            
            # Fallback final con S3Service
            if self.s3_service:
                logger.info("🔄 Último intento con S3Service...")
                return self._descargar_con_s3service_como_fallback(url_s3)
            
            return None
    
    def _descargar_con_s3service_como_fallback(self, url_s3: str) -> Optional[bytes]:
        """
        Fallback usando S3Service cuando la URL directa falla
        Extrae el key de S3 de la URL y usa S3Service para descargar
        """
        try:
            # Extraer el S3 key de la URL
            # URL ejemplo: https://mesadeayudacqs3img.s3.us-east-2.amazonaws.com/ots/ot_1945_20251009_213742_0fd60558_Captura_de_pantalla_2024-01-15_104827.png
            if 'amazonaws.com/' in url_s3:
                s3_key = url_s3.split('amazonaws.com/')[1]
                logger.info(f"🔑 S3 Key extraído: {s3_key}")
                
                # Intentar descarga con S3Service
                download_result = self.s3_service.download_file_from_s3(s3_key)
                if download_result.get('success', False):
                    content = download_result.get('content')
                    if content:
                        logger.info(f"✅ Imagen descargada con S3Service usando key: {s3_key} ({len(content)} bytes)")
                        return content
                
                # Si no funciona con el key directo, generar URL presignada
                logger.info("🔄 Intentando con URL presignada...")
                url_result = self.s3_service.generate_presigned_download_url(s3_key)
                if url_result.get('success', False):
                    presigned_url = url_result.get('url')
                    if presigned_url:
                        response = requests.get(presigned_url, timeout=30)
                        response.raise_for_status()
                        logger.info(f"✅ Imagen descargada con URL presignada ({len(response.content)} bytes)")
                        return response.content
            
            logger.warning(f"❌ No se pudo descargar con S3Service desde: {url_s3}")
            return None
            
        except Exception as e:
            logger.error(f"❌ Error en fallback S3Service: {str(e)}")
            return None
    
    def _debug_listar_archivos_s3(self, nombre_buscado: str):
        """
        Función de debug para listar archivos en S3 y ayudar a encontrar el nombre correcto
        """
        try:
            # Extraer prefijo del nombre del archivo para buscar similares
            prefijo_busqueda = nombre_buscado.split('_')[0] if '_' in nombre_buscado else nombre_buscado[:10]
            
            logger.info(f"� DEBUG: Listando archivos S3 que contengan '{prefijo_busqueda}'")
            
            # Listar archivos con diferentes prefijos
            folders_to_check = ['', 'images/', 'uploads/', 'b2c/']
            
            for folder in folders_to_check:
                try:
                    list_result = self.s3_service.list_files(folder=folder.rstrip('/') if folder else 'images', limit=50)
                    if list_result.get('success', False):
                        files = list_result.get('files', [])
                        matching_files = [f for f in files if prefijo_busqueda.lower() in f.get('key', '').lower()]
                        
                        if matching_files:
                            logger.info(f"📁 Archivos encontrados en '{folder}':")
                            for file_info in matching_files[:5]:  # Mostrar solo los primeros 5
                                logger.info(f"  - {file_info.get('key', 'Sin nombre')}")
                        else:
                            logger.debug(f"📁 No se encontraron archivos similares en '{folder}'")
                except Exception as e:
                    logger.debug(f"Error listando carpeta '{folder}': {str(e)}")
                    
        except Exception as e:
            logger.error(f"Error en debug de listado S3: {str(e)}")
    
    def _redimensionar_imagen_optimizada(self, imagen_bytes: bytes, max_width: int, max_height: int, layout_config: dict = None) -> Optional[bytes]:
        """
        Redimensionar imagen con optimización avanzada para área Excel A40:H42
        Considerando el layout específico y orientación de la imagen
        """
        try:
            # Abrir imagen con PIL
            with PILImage.open(io.BytesIO(imagen_bytes)) as img:
                # Convertir a RGB si es necesario (para JPEGs)
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                
                # Obtener dimensiones originales
                original_width, original_height = img.size
                
                # 📏 ANÁLISIS AVANZADO DE ORIENTACIÓN
                aspect_ratio = original_width / original_height
                if aspect_ratio < 0.6:
                    orientation = "muy_vertical"  # Imagen muy alta (ej: captura móvil)
                elif aspect_ratio < 0.8:
                    orientation = "vertical"     # Imagen alta
                elif aspect_ratio > 1.8:
                    orientation = "muy_horizontal"  # Imagen muy ancha (ej: panorámica)
                elif aspect_ratio > 1.2:
                    orientation = "horizontal"   # Imagen ancha
                else:
                    orientation = "cuadrada"     # Imagen cuadrada o casi cuadrada
                
                logger.info(f"📐 Imagen {original_width}x{original_height} - Orientación: {orientation} (ratio: {aspect_ratio:.2f})")
                
                # 🎯 CÁLCULO OPTIMIZADO SEGÚN LAYOUT Y ORIENTACIÓN
                layout_type = layout_config.get("layout_type", "unknown") if layout_config else "unknown"
                
                if layout_type == "single_horizontal" and orientation in ["horizontal", "muy_horizontal"]:
                    # Imagen horizontal única: maximizar ancho, controlar altura
                    target_width = min(max_width, original_width)  # No exceder ancho máximo
                    target_height = min(max_height * 0.8, int(target_width / aspect_ratio))  # 80% del alto disponible
                elif layout_type == "single_vertical" and orientation in ["vertical", "muy_vertical"]:
                    # Imagen vertical única: maximizar altura, controlar ancho
                    target_height = min(max_height, original_height)  # No exceder altura máxima
                    target_width = min(max_width * 0.7, int(target_height * aspect_ratio))  # 70% del ancho disponible
                elif orientation == "muy_vertical":
                    # Imagen muy alta: priorizar que quepa en altura
                    target_height = max_height
                    target_width = min(max_width, int(target_height * aspect_ratio))
                elif orientation == "muy_horizontal":
                    # Imagen muy ancha: priorizar que quepa en ancho
                    target_width = max_width
                    target_height = min(max_height, int(target_width / aspect_ratio))
                else:
                    # Cálculo estándar manteniendo proporción
                    ratio = min(max_width/original_width, max_height/original_height)
                    target_width = int(original_width * ratio)
                    target_height = int(original_height * ratio)
                
                # 🔧 APLICAR REDIMENSIONAMIENTO SI ES NECESARIO
                if target_width < original_width or target_height < original_height:
                    img = img.resize((target_width, target_height), PILImage.Resampling.LANCZOS)
                    logger.info(f"🔧 Redimensionamiento optimizado aplicado: {target_width}x{target_height}")
                else:
                    target_width, target_height = original_width, original_height
                    logger.info(f"📏 Imagen conserva tamaño original (ya cabe en área disponible)")
                
                # Convertir de vuelta a bytes
                output = io.BytesIO()
                img.save(output, format='PNG', quality=95, optimize=True)
                output.seek(0)
                
                logger.info(f"✅ Imagen procesada: {original_width}x{original_height} → {target_width}x{target_height} ({orientation})")
                
                return output.getvalue()
                
        except Exception as e:
            logger.error(f"❌ Error redimensionando imagen optimizada: {str(e)}")
            return None
    
    def _redimensionar_imagen(self, imagen_bytes: bytes, max_width: int, max_height: int) -> Optional[bytes]:
        """
        Función legacy - usar _redimensionar_imagen_optimizada para mejor rendimiento
        """
        return self._redimensionar_imagen_optimizada(imagen_bytes, max_width, max_height)
    
    def _llenar_celdas_combinadas(self, ws, datos_ot: Dict[str, Any], folio: int):
        """
        Manejar celdas combinadas que requieren tratamiento especial
        NOTA: Esta función ya no es necesaria porque la nueva función _llenar_campos_excel
        maneja las celdas combinadas automáticamente
        """
        logger.info("Función de celdas combinadas llamada - manejo integrado en _llenar_campos_excel")
        pass
    
    def _convertir_con_libreoffice(self, excel_path: Path, output_dir: Path) -> Path:
        """
        Convertir Excel a PDF usando LibreOffice headless manteniendo formato exacto
        
        Args:
            excel_path: Archivo Excel a convertir
            output_dir: Directorio de salida
            
        Returns:
            Path: Ruta al PDF generado
        """
        logger.info("Convirtiendo a PDF con LibreOffice headless...")
        
        # Detectar comando LibreOffice según el sistema
        possible_commands = [
            'soffice',  # Linux/macOS
            'libreoffice',  # Linux alternativo
            '"C:\\Program Files\\LibreOffice\\program\\soffice.exe"',  # Windows estándar
            '"C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe"',  # Windows x86
        ]
        
        soffice_cmd = None
        for cmd in possible_commands:
            try:
                test_result = subprocess.run([cmd.strip('"'), '--version'], 
                                           capture_output=True, text=True, timeout=5)
                if test_result.returncode == 0:
                    soffice_cmd = cmd.strip('"')
                    logger.info(f"LibreOffice encontrado: {soffice_cmd}")
                    break
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue
        
        if not soffice_cmd:
            logger.error("LibreOffice no encontrado en rutas estándar")
            raise RuntimeError("LibreOffice no encontrado en el sistema")
        
        # Comando LibreOffice headless para conversión exacta
        cmd = [
            soffice_cmd,
            '--headless',
            '--convert-to', 'pdf:calc_pdf_Export',
            '--outdir', str(output_dir),
            str(excel_path)
        ]
        
        try:
            # Ejecutar conversión
            logger.info(f"Ejecutando: {' '.join(cmd)}")
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            if result.returncode != 0:
                logger.error(f"Error en LibreOffice: {result.stderr}")
                raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")
            
            # Buscar archivo PDF generado
            pdf_name = excel_path.stem + '.pdf'
            pdf_path = output_dir / pdf_name
            
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF no generado: {pdf_path}")
            
            logger.info(f"PDF convertido exitosamente: {pdf_path}")
            return pdf_path
            
        except subprocess.TimeoutExpired:
            logger.error("LibreOffice conversion timeout")
            raise RuntimeError("LibreOffice conversion timeout")
        except FileNotFoundError:
            logger.error("LibreOffice no encontrado. Instalar LibreOffice.")
            raise RuntimeError("LibreOffice no encontrado en el sistema")
    
    def verificar_libreoffice(self) -> bool:
        """
        Verificar si LibreOffice está disponible en el sistema
        
        Returns:
            bool: True si LibreOffice está disponible
        """
        # Posibles comandos de LibreOffice
        possible_commands = [
            'soffice',  # Linux/macOS
            'libreoffice',  # Linux alternativo
            '/usr/bin/soffice',  # Ruta absoluta Linux
            '/usr/bin/libreoffice',  # Ruta absoluta Linux alternativa
            'C:\\Program Files\\LibreOffice\\program\\soffice.exe',  # Windows estándar
            'C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe',  # Windows x86
        ]
        
        for cmd in possible_commands:
            try:
                logger.info(f"🔍 Probando comando LibreOffice: {cmd}")
                result = subprocess.run([cmd, '--version'], capture_output=True, text=True, timeout=10)
                logger.info(f"🔍 Return code: {result.returncode}")
                logger.info(f"🔍 Stdout: {result.stdout.strip()}")
                logger.info(f"🔍 Stderr: {result.stderr.strip()}")
                
                if result.returncode == 0:
                    logger.info(f"✅ LibreOffice disponible: {cmd}")
                    logger.info(f"✅ Versión: {result.stdout.strip()}")
                    return True
                else:
                    logger.warning(f"⚠️ Comando falló: {cmd} - Return code: {result.returncode}")
                    
            except FileNotFoundError as e:
                logger.warning(f"⚠️ Comando no encontrado: {cmd} - {e}")
                continue
            except subprocess.TimeoutExpired as e:
                logger.warning(f"⚠️ Timeout en comando: {cmd} - {e}")
                continue
            except Exception as e:
                logger.warning(f"⚠️ Error inesperado con comando: {cmd} - {e}")
                continue
        
        logger.error("❌ LibreOffice no encontrado en ninguna ruta estándar")
        return False
    
    def _convertir_con_excel_com(self, excel_path: Path, output_dir: Path) -> Path:
        """
        Convertir Excel a PDF usando Excel COM (solo Windows, para desarrollo)
        
        Args:
            excel_path: Archivo Excel a convertir
            output_dir: Directorio de salida
            
        Returns:
            Path: Ruta al PDF generado
        """
        if os.name != 'nt':
            raise RuntimeError("Excel COM solo disponible en Windows")
        
        logger.info("Convirtiendo a PDF con Excel COM (Windows)...")
        
        try:
            import win32com.client
        except ImportError:
            logger.error("pywin32 no instalado. Instalar con: pip install pywin32")
            raise RuntimeError("pywin32 requerido para Excel COM")
        
        excel_app = None
        workbook = None
        try:
            # Crear aplicación Excel
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False
            
            # Abrir workbook
            workbook = excel_app.Workbooks.Open(str(excel_path.absolute()))
            
            # PDF path
            pdf_path = output_dir / f"{excel_path.stem}.pdf"
            
            # Exportar como PDF manteniendo formato
            workbook.ExportAsFixedFormat(
                Type=0,  # xlTypePDF
                Filename=str(pdf_path.absolute()),
                Quality=0,  # xlQualityStandard
                IncludeDocProperties=True,  # Parámetro correcto
                IgnorePrintAreas=False,
                From=1,
                To=1,
                OpenAfterPublish=False
            )
            
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF no generado: {pdf_path}")
            
            logger.info(f"PDF convertido con Excel COM: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            logger.error(f"Error en Excel COM: {e}")
            raise RuntimeError(f"Excel COM conversion failed: {e}")
        finally:
            # Cerrar workbook y Excel de forma segura
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
            except:
                pass
            try:
                if excel_app:
                    excel_app.Quit()
            except:
                pass
            
            # Esperar un poco para que Excel libere los archivos
            import time
            time.sleep(0.5)
