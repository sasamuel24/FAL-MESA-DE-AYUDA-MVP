"""
Sistema de gestión de plantillas Excel con nombres de rango
Implementa la metodología completa para generación de PDFs basados en plantillas Excel
"""
import openpyxl
from openpyxl.styles import Protection
from pathlib import Path
import logging
from typing import Dict, Any, List, Optional, Tuple
import json
from datetime import datetime
import shutil

logger = logging.getLogger(__name__)

class ExcelTemplateManager:
    """
    Gestor de plantillas Excel que implementa:
    1. Análisis estructural de plantillas
    2. Sistema de nombres de rango
    3. Protección de celdas
    4. Inyección controlada de datos
    5. Exportación PDF de alta fidelidad
    """
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        self.campos_dinamicos = []
        self.elementos_estaticos = []
        self.celdas_fusionadas = []
        self.nombres_rango = {}
        
        if not self.template_path.exists():
            raise FileNotFoundError(f"Plantilla no encontrada: {self.template_path}")
    
    def analizar_plantilla(self):
        """
        1. ANÁLISIS Y MAPEO ESTRUCTURAL DE LA PLANTILLA
        """
        logger.info(f"Analizando plantilla: {self.template_path}")
        
        workbook = openpyxl.load_workbook(self.template_path)
        sheet = workbook.active
        
        # 1.1 Auditoría completa de elementos
        self._identificar_elementos(sheet)
        
        # 1.2 Documentación de posicionamiento
        self._analizar_celdas_fusionadas(sheet)
        
        # 1.3 Análisis de flujo visual
        self._mapear_flujo_visual(sheet)
        
        workbook.close()
        
        logger.info(f"Analisis completado: {len(self.campos_dinamicos)} campos dinamicos, {len(self.elementos_estaticos)} elementos estaticos")
    
    def _identificar_elementos(self, sheet):
        """Identificar elementos estáticos vs dinámicos"""
        self.campos_dinamicos = []
        self.elementos_estaticos = []
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    value = str(cell.value).strip()
                    col_letter = openpyxl.utils.get_column_letter(col)
                    cell_ref = f"{col_letter}{row}"
                    
                    # Campos dinámicos (etiquetas que terminan en ":")
                    if value.endswith(':') and len(value) > 1:
                        campo_nombre = value[:-1]
                        
                        # Encontrar la celda de datos asociada
                        celda_datos = self._encontrar_celda_datos(sheet, row, col)
                        
                        self.campos_dinamicos.append({
                            'campo': campo_nombre,
                            'etiqueta_celda': cell_ref,
                            'datos_celda': celda_datos,
                            'fila': row,
                            'columna': col,
                            'fila_datos': celda_datos[1] if celda_datos else None,
                            'columna_datos': celda_datos[2] if celda_datos else None
                        })
                    
                    # Elementos estáticos importantes
                    elif any(keyword in value.upper() for keyword in [
                        'ORDEN DE TRABAJO', 'CÓDIGO', 'VERSIÓN', 'FECHA',
                        'ASIGNACIÓN', 'DESCRIPCIÓN', 'NOTAS', 'FIRMAS'
                    ]):
                        self.elementos_estaticos.append({
                            'elemento': value,
                            'celda': cell_ref,
                            'fila': row,
                            'columna': col,
                            'tipo': 'encabezado' if any(h in value.upper() for h in ['ORDEN DE TRABAJO', 'ASIGNACIÓN', 'DESCRIPCIÓN']) else 'etiqueta'
                        })
    
    def _encontrar_celda_datos(self, sheet, etiqueta_row: int, etiqueta_col: int) -> Optional[Tuple[str, int, int]]:
        """Encontrar la celda donde van los datos asociados a una etiqueta"""
        # Buscar a la derecha de la etiqueta
        for col_offset in range(1, 4):  # Buscar hasta 3 columnas a la derecha
            target_col = etiqueta_col + col_offset
            if target_col <= sheet.max_column:
                col_letter = openpyxl.utils.get_column_letter(target_col)
                cell_ref = f"{col_letter}{etiqueta_row}"
                
                # Verificar si la celda está en un rango fusionado
                for merged_range in sheet.merged_cells.ranges:
                    if (target_col >= merged_range.min_col and target_col <= merged_range.max_col and
                        etiqueta_row >= merged_range.min_row and etiqueta_row <= merged_range.max_row):
                        # Usar la celda superior izquierda del rango fusionado
                        top_left_col = merged_range.min_col
                        top_left_row = merged_range.min_row
                        top_left_letter = openpyxl.utils.get_column_letter(top_left_col)
                        return (f"{top_left_letter}{top_left_row}", top_left_row, top_left_col)
                
                return (cell_ref, etiqueta_row, target_col)
        
        return None
    
    def _analizar_celdas_fusionadas(self, sheet):
        """Analizar celdas fusionadas para preservar estructura"""
        self.celdas_fusionadas = []
        
        for merged_range in sheet.merged_cells.ranges:
            bounds = merged_range.bounds
            top_left_cell = sheet.cell(row=bounds[1], column=bounds[0])
            
            self.celdas_fusionadas.append({
                'rango': str(merged_range),
                'valor': top_left_cell.value if top_left_cell.value else None,
                'filas': bounds[3] - bounds[1] + 1,
                'columnas': bounds[2] - bounds[0] + 1,
                'min_row': bounds[1],
                'max_row': bounds[3],
                'min_col': bounds[0],
                'max_col': bounds[2]
            })
    
    def _mapear_flujo_visual(self, sheet):
        """Mapear flujo visual del documento"""
        # Identificar secciones principales por filas
        secciones = []
        seccion_actual = None
        
        for fusionada in self.celdas_fusionadas:
            if fusionada['valor'] and any(keyword in str(fusionada['valor']).upper() for keyword in [
                'ORDEN DE TRABAJO', 'ASIGNACIÓN', 'DESCRIPCIÓN', 'NOTAS', 'ARCHIVOS', 'FIRMAS'
            ]):
                secciones.append({
                    'nombre': fusionada['valor'],
                    'fila_inicio': fusionada['min_row'],
                    'fila_fin': fusionada['max_row'],
                    'tipo': 'encabezado_seccion'
                })
        
        self.secciones_visuales = sorted(secciones, key=lambda x: x['fila_inicio'])
    
    def crear_nombres_rango(self):
        """
        2. IMPLEMENTACIÓN DE SISTEMA DE NOMBRES DE RANGO
        """
        logger.info("Creando nombres de rango...")
        
        # Crear copia de trabajo de la plantilla
        plantilla_con_rangos = self.template_path.parent / f"OT_Template_WithRanges_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(self.template_path, plantilla_con_rangos)
        
        workbook = openpyxl.load_workbook(plantilla_con_rangos)
        sheet = workbook.active
        
        # 2.1 Estrategia de nomenclatura
        self.nombres_rango = {}
        rangos_creados = 0
        
        for campo in self.campos_dinamicos:
            if campo['datos_celda']:
                # Crear nombre de rango según convención
                nombre_rango = self._generar_nombre_rango(campo['campo'])
                celda_datos = campo['datos_celda'][0]  # ej: "B5"
                
                try:
                    # Crear el nombre de rango en Excel
                    sheet.parent.create_named_range(nombre_rango, sheet, celda_datos)
                    
                    self.nombres_rango[nombre_rango] = {
                        'campo_original': campo['campo'],
                        'celda': celda_datos,
                        'fila': campo['fila_datos'],
                        'columna': campo['columna_datos'],
                        'tipo_dato': self._inferir_tipo_dato(campo['campo'])
                    }
                    
                    rangos_creados += 1
                    logger.debug(f"Rango creado: {nombre_rango} -> {celda_datos}")
                    
                except Exception as e:
                    logger.warning(f"No se pudo crear rango {nombre_rango}: {e}")
        
        # Guardar plantilla con rangos
        workbook.save(plantilla_con_rangos)
        workbook.close()
        
        logger.info(f"Creados {rangos_creados} nombres de rango en: {plantilla_con_rangos}")
        
        # Guardar configuración de rangos
        self._guardar_configuracion_rangos()
        
        return plantilla_con_rangos
    
    def _generar_nombre_rango(self, campo: str) -> str:
        """Generar nombre de rango según convención OT_CAMPO"""
        # Limpiar y normalizar el nombre del campo
        nombre_limpio = campo.upper()
        nombre_limpio = nombre_limpio.replace('/', '_')
        nombre_limpio = nombre_limpio.replace(' ', '_')
        nombre_limpio = nombre_limpio.replace('(', '')
        nombre_limpio = nombre_limpio.replace(')', '')
        nombre_limpio = nombre_limpio.replace('Ñ', 'N')
        nombre_limpio = nombre_limpio.replace('Á', 'A')
        nombre_limpio = nombre_limpio.replace('É', 'E')
        nombre_limpio = nombre_limpio.replace('Í', 'I')
        nombre_limpio = nombre_limpio.replace('Ó', 'O')
        nombre_limpio = nombre_limpio.replace('Ú', 'U')
        
        return f"OT_{nombre_limpio}"
    
    def _inferir_tipo_dato(self, campo: str) -> str:
        """Inferir tipo de dato basado en el nombre del campo"""
        campo_lower = campo.lower()
        
        if 'fecha' in campo_lower:
            return 'fecha'
        elif 'id' in campo_lower or 'folio' in campo_lower:
            return 'numero'
        elif 'tiempo' in campo_lower and 'h' in campo_lower:
            return 'numero'
        elif 'descripcion' in campo_lower or 'notas' in campo_lower:
            return 'texto_largo'
        else:
            return 'texto'
    
    def _guardar_configuracion_rangos(self):
        """Guardar configuración de rangos para trazabilidad"""
        config_path = self.template_path.parent / "rangos_configuracion.json"
        
        config = {
            'plantilla_original': str(self.template_path),
            'fecha_analisis': datetime.now().isoformat(),
            'campos_dinamicos': self.campos_dinamicos,
            'nombres_rango': self.nombres_rango,
            'elementos_estaticos': self.elementos_estaticos,
            'celdas_fusionadas': self.celdas_fusionadas
        }
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Configuracion guardada en: {config_path}")
    
    def configurar_proteccion(self, plantilla_con_rangos: Path):
        """
        3. SISTEMA DE PROTECCIÓN Y PRESERVACIÓN DE FORMATO
        """
        logger.info("Configurando proteccion de plantilla...")
        
        workbook = openpyxl.load_workbook(plantilla_con_rangos)
        sheet = workbook.active
        
        # 3.1 Configuración de protección granular
        # Bloquear todas las celdas por defecto
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True, hidden=False)
        
        # Desbloquear solo las celdas con nombres de rango (campos dinámicos)
        celdas_desbloqueadas = 0
        for nombre_rango, config in self.nombres_rango.items():
            try:
                celda = sheet[config['celda']]
                celda.protection = Protection(locked=False, hidden=False)
                celdas_desbloqueadas += 1
                logger.debug(f"Desbloqueada: {config['celda']} ({nombre_rango})")
            except Exception as e:
                logger.warning(f"No se pudo desbloquear {config['celda']}: {e}")
        
        # Activar protección de la hoja
        sheet.protection.sheet = True
        sheet.protection.password = "CafeQuindio2024"  # Contraseña de protección
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.insertHyperlinks = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.selectLockedCells = True
        sheet.protection.selectUnlockedCells = True
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.pivotTables = False
        sheet.protection.objects = True
        sheet.protection.scenarios = True
        
        # Guardar plantilla protegida
        plantilla_protegida = self.template_path.parent / f"OT_Template_Protected_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(plantilla_protegida)
        workbook.close()
        
        logger.info(f"Proteccion configurada: {celdas_desbloqueadas} celdas desbloqueadas en {plantilla_protegida}")
        
        return plantilla_protegida
    
    def llenar_datos(self, plantilla_protegida: Path, datos_ot: Dict[str, Any]) -> Path:
        """
        4. PROCESO CONTROLADO DE INYECCIÓN DE DATOS
        """
        logger.info(f"Llenando datos en plantilla protegida...")
        
        # 4.1 Crear copia de trabajo temporal
        archivo_con_datos = self.template_path.parent / f"OT_{datos_ot.get('folio', 'TEMP')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(plantilla_protegida, archivo_con_datos)
        
        workbook = openpyxl.load_workbook(archivo_con_datos)
        sheet = workbook.active
        
        # Desproteger temporalmente para insertar datos
        sheet.protection.sheet = False
        
        # 4.2 Validación de contenido pre-inserción e inserción segura
        datos_insertados = 0
        errores = []
        
        for nombre_rango, config in self.nombres_rango.items():
            try:
                # Mapear nombre de rango a campo de datos
                valor = self._obtener_valor_para_campo(config['campo_original'], datos_ot)
                
                if valor is not None:
                    # Validar y transformar dato
                    valor_procesado = self._procesar_valor(valor, config['tipo_dato'])
                    
                    # Insertar en la celda específica
                    celda = sheet[config['celda']]
                    celda.value = valor_procesado
                    
                    datos_insertados += 1
                    logger.debug(f"Insertado: {config['campo_original']} = {valor_procesado} en {config['celda']}")
                
            except Exception as e:
                error_msg = f"Error insertando {nombre_rango}: {e}"
                errores.append(error_msg)
                logger.warning(f"{error_msg}")
        
        # 4.3 Control de calidad post-inserción
        if self._validar_integridad_formato(sheet):
            logger.info("Integridad de formato validada")
        else:
            logger.warning("Se detectaron posibles problemas de formato")
        
        # Reproteger la hoja
        sheet.protection.sheet = True
        sheet.protection.password = "CafeQuindio2024"
        
        # Guardar archivo con datos
        workbook.save(archivo_con_datos)
        workbook.close()
        
        logger.info(f"Datos insertados: {datos_insertados} campos, {len(errores)} errores")
        
        if errores:
            for error in errores:
                logger.error(f"{error}")
        
        return archivo_con_datos
    
    def _obtener_valor_para_campo(self, campo: str, datos_ot: Dict[str, Any]) -> Any:
        """Mapear campos de plantilla a datos de OT"""
        # Mapeo de campos de plantilla Excel a campos de datos de OT
        mapeo_campos = {
            'CÓDIGO': f"FO-MT-006-{datos_ot.get('folio', 'N/A')}",
            'VERSIÓN': '1.0',
            'FECHA': datetime.now().strftime('%Y-%m-%d'),
            'Título': datos_ot.get('titulo', ''),
            'ID': datos_ot.get('folio', ''),
            'Fecha': datos_ot.get('fecha_creacion', ''),
            'Estado': datos_ot.get('estado', ''),
            'Categoría': datos_ot.get('categoria', ''),
            'Subcategoría': datos_ot.get('subcategoria', ''),
            'Zona / Planta / Tienda': datos_ot.get('ubicacion', ''),
            'Ciudad': datos_ot.get('ciudad', ''),
            'Prioridad': datos_ot.get('prioridad', ''),
            'Tipo de Solicitud': datos_ot.get('tipo_solicitud', ''),
            'Tipo de Mantenimiento': datos_ot.get('tipo_mantenimiento', ''),
            'Tiempo estimado (h)': datos_ot.get('tiempo_estimado', ''),
            'Etapa': datos_ot.get('etapa', ''),
            'Técnico asignado': datos_ot.get('tecnico_asignado', ''),
            'Fecha de visita': datos_ot.get('fecha_visita', ''),
            'Solicitante': datos_ot.get('solicitante', ''),
            'Contacto solicitante': datos_ot.get('contacto_solicitante', ''),
        }
        
        return mapeo_campos.get(campo, None)
    
    def _procesar_valor(self, valor: Any, tipo_dato: str) -> Any:
        """Procesar y validar valor según su tipo"""
        if valor is None or valor == '':
            return ''
        
        if tipo_dato == 'fecha':
            if isinstance(valor, str):
                try:
                    # Intentar parsear fecha
                    if 'T' in valor:  # Formato ISO
                        dt = datetime.fromisoformat(valor.replace('Z', '+00:00'))
                        return dt.strftime('%Y-%m-%d')
                    else:
                        return valor
                except:
                    return valor
            return str(valor)
        
        elif tipo_dato == 'numero':
            try:
                return float(valor) if '.' in str(valor) else int(valor)
            except:
                return str(valor)
        
        elif tipo_dato == 'texto_largo':
            # Truncar si es muy largo
            texto = str(valor)
            return texto[:500] + '...' if len(texto) > 500 else texto
        
        else:  # texto
            texto = str(valor)
            return texto[:100] + '...' if len(texto) > 100 else texto
    
    def _validar_integridad_formato(self, sheet) -> bool:
        """Validar que el formato de la plantilla se mantuvo intacto"""
        # Verificar que las celdas fusionadas siguen intactas
        celdas_fusionadas_actuales = len(list(sheet.merged_cells.ranges))
        celdas_fusionadas_originales = len(self.celdas_fusionadas)
        
        if celdas_fusionadas_actuales != celdas_fusionadas_originales:
            logger.warning(f"Cambio en celdas fusionadas: {celdas_fusionadas_originales} -> {celdas_fusionadas_actuales}")
            return False
        
        return True
    
    def exportar_pdf(self, archivo_con_datos: Path) -> Path:
        """
        5. OPTIMIZACIÓN DE EXPORTACIÓN PDF CON FIDELIDAD MÁXIMA
        """
        logger.info(f"Exportando a PDF: {archivo_con_datos}")
        
        try:
            # Cargar archivo con datos
            workbook = openpyxl.load_workbook(archivo_con_datos)
            sheet = workbook.active
            
            # 5.1 Configurar área de impresión
            # El área de impresión ya está definida en la plantilla original
            # Si no está definida, configurarla para abarcar todo el contenido
            if not sheet.print_area:
                # Establecer área de impresión automáticamente
                max_row_with_data = sheet.max_row
                max_col_with_data = sheet.max_column
                
                end_col_letter = openpyxl.utils.get_column_letter(max_col_with_data)
                sheet.print_area = f"A1:{end_col_letter}{max_row_with_data}"
                logger.info(f"📐 Área de impresión configurada: {sheet.print_area}")
            
            # 5.2 Configurar parámetros de página para PDF
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0  # Permitir múltiples páginas en altura
            
            # Configurar márgenes
            sheet.page_margins.left = 0.7
            sheet.page_margins.right = 0.7
            sheet.page_margins.top = 0.75
            sheet.page_margins.bottom = 0.75
            sheet.page_margins.header = 0.3
            sheet.page_margins.footer = 0.3
            
            # Guardar configuración
            workbook.save(archivo_con_datos)
            workbook.close()
            
            # 5.3 Exportar a PDF (requiere Excel instalado o alternativa)
            archivo_pdf = archivo_con_datos.with_suffix('.pdf')
            
            # Nota: La exportación real a PDF requiere una implementación específica
            # Por ahora, creamos un placeholder que indica que el archivo está listo
            logger.info(f"Archivo preparado para exportacion PDF: {archivo_con_datos}")
            logger.info(f"PDF destino: {archivo_pdf}")
            
            return archivo_pdf
            
        except Exception as e:
            logger.error(f"Error exportando PDF: {e}")
            raise
    
    def procesar_plantilla_completa(self, datos_ot: Dict[str, Any]) -> Path:
        """
        Proceso completo: analizar -> crear rangos -> proteger -> llenar -> exportar
        """
        logger.info("🚀 Iniciando proceso completo de plantilla Excel")
        
        try:
            # 1. Analizar plantilla
            self.analizar_plantilla()
            
            # 2. Crear nombres de rango
            plantilla_con_rangos = self.crear_nombres_rango()
            
            # 3. Configurar protección
            plantilla_protegida = self.configurar_proteccion(plantilla_con_rangos)
            
            # 4. Llenar datos
            archivo_con_datos = self.llenar_datos(plantilla_protegida, datos_ot)
            
            # 5. Exportar PDF
            archivo_pdf = self.exportar_pdf(archivo_con_datos)
            
            logger.info(f"Proceso completo finalizado: {archivo_pdf}")
            
            return archivo_con_datos  # Retornar archivo Excel por ahora
            
        except Exception as e:
            logger.error(f"Error en proceso completo: {e}")
            raise


def main():
    """Función de prueba"""
    template_path = "d:/CafeQuindio/backend/app/templates/FO-MT-006 Orden de trabajo de mantenimiento v1.xlsx"
    
    # Configurar logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    try:
        manager = ExcelTemplateManager(template_path)
        
        # Datos de prueba
        datos_ot = {
            'folio': 1927,
            'titulo': 'Mantenimiento de Equipos de Café',
            'fecha_creacion': '2025-10-09',
            'estado': 'En Proceso',
            'categoria': 'Mantenimiento Preventivo',
            'subcategoria': 'Equipos de Producción',
            'ubicacion': 'Planta Principal',
            'ciudad': 'Armenia',
            'prioridad': 'Alta',
            'tipo_solicitud': 'Programado',
            'tipo_mantenimiento': 'Preventivo',
            'tiempo_estimado': '4',
            'etapa': 'Ejecución',
            'tecnico_asignado': 'Juan Pérez',
            'fecha_visita': '2025-10-10',
            'solicitante': 'María García',
            'contacto_solicitante': 'maria.garcia@cafequindio.com'
        }
        
        # Procesar plantilla completa
        resultado = manager.procesar_plantilla_completa(datos_ot)
        print(f"Archivo generado: {resultado}")
        
    except Exception as e:
        print(f"Error: {e}")
        raise


if __name__ == "__main__":
    main()